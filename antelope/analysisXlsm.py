# -*- coding: utf-8 -*-
"""
分析 Amazon Flat File 模板 (.xlsm)，只读不写，结果以 JSON 格式输出。

用途：解析模板文件中「表头」以及该表头（字段）单元格的可选值，以 JSON 格式给出，
供人工/程序填写时参考。

数据来源（默认 .xlsm_template_base，即需求里的基础模板 B）：
  - 主模板工作表（fr=Modèle / de=Vorlage，按国家配置 SHEET_NAMES 取）：
                                row1 的 settings=... 串里带有 labelRow / attributeRow / dataRow 参数；
                                labelRow 是表头（人类可读列名），attributeRow 是属性键，dataRow 是数据起始行。
  - 字段枚举表工作表（fr=Valeurs valides / de=Gültige Werte）：
                                第 1列为分组标题；第 2 列为字段名（形如 "字段名 - [ COAT ]"）；
                                第 3 列起横向排列该字段的可选值。

匹配方式：以主模板的表头文本（归一化：统一 \xa0/空白、去首尾空白）去匹配
字段枚举表中字段名（去掉 " - [ COAT ]" 后缀后同样归一化）；命中则列出其可选值，
否则该列为自由文本/无预定义枚举。

按 11409 需求的分工（A/B/C/M/D 角色见 zconfig.constant.py）：
  - 对 基础模板 B 解析  → blank.json（「已填列」，column_diff 的减数）
  - 对 完整模板 C 解析  → completed.json（「完整列」，column_diff 的被减数）
  - 生成 completed 示例：
      python analysisXlsm.py xlsm/.xlsm_template_complete -o intermediate/fr_shirt/fr_shirt_completed.json

用法：
    python analysisXlsm.py [文件路径] [--output 输出.json] [--sheets 工作表1 工作表2]
                           [--max-values N] [--show-data]
示例：
    python analysisXlsm.py                      # 默认解析基础模板 B，输出 blank.json
    python analysisXlsm.py xlsm/.xlsm_template_complete -o intermediate/fr_shirt/fr_shirt_completed.json

说明：JSON 中 columns 数组的每个元素（每列）预留了 reserve_flag / reserve_mark 两个标识位字段，
当前恒为 null，仅供后续程序复用标记，不影响其余解析结果。
"""

import argparse
import json
import os
import re
import sys

import openpyxl

from common import setup_log, zcfg

# 默认输入模板文件（基础模板 B，解析得到 blank.json / 「已填列」）
DEFAULT_INPUT_PATH = zcfg.TEMPLATE_B

# 默认输出 JSON（模板分析结果）
DEFAULT_OUTPUT_PATH = zcfg.CFG_INTERMEDIATE["blank_json"]

# 默认参与解析的工作表
DEFAULT_SHEETS = list(zcfg.CFG_RUN["default_sheets"])


# --------------------------------------------------------------------------- #
# 文本归一化：让不同来源的法文标签能可靠地相互匹配
# --------------------------------------------------------------------------- #
def normalize(text: str) -> str:
    """统一空白类字符、大小写与标点半角化，用于表头与字段名的匹配。"""
    if text is None:
        return ""
    s = str(text)
    # 把所有不可见空白（普通空格、\xa0、全角空格等）统一为普通空格
    s = re.sub(r"[\s\u00a0\u2007\u202f]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = s.casefold()  # 不区分大小写
    # 把全角/弯引号归一为 ASCII 引号，减少差异
    for a, b in (("\u2018", "'"), ("\u2019", "'"), ("\u201a", "'"),
                 ("\u201c", '"'), ("\u201d", '"'), ("\u201e", '"'),
                 ("\u200b", "")):
        s = s.replace(a, b)
    return s.strip()


def field_key(field_name: str) -> str:
    """把字段枚举表中字段名去掉 ' - [XXXX]' 后缀后做归一化。"""
    s = re.sub(r"\s*-\s*\[.*?\]\s*$", "", str(field_name))
    return normalize(s)


# --------------------------------------------------------------------------- #
# 解析字段枚举表：分组标题 + 字段 -> 可选值列表
# --------------------------------------------------------------------------- #
def parse_valid_values(ws) -> tuple:
    """
    遍历字段枚举表工作表。

    返回:
        valid_values: {归一化后的字段名: [(原始字段名, [可选值, ...]), ...]}
                      同一字段名可能出现多次（如两个 "Style de col"：
                      22 值的 collar_style 与 31 值的 neck），按出现顺序保留全部；
        groups:       [分组标题, ...]（按出现顺序，去重）
    """
    valid_values = {}
    groups = []
    seen = set()

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        col1 = row[0] if len(row) > 0 else None  # 分组标题
        col2 = row[1] if len(row) > 1 else None  # 字段名
        if col1 is not None and str(col1).strip():
            group = str(col1).strip()
            if group not in seen:
                seen.add(group)
                groups.append(group)
        if col2 is not None and str(col2).strip():
            field_raw = str(col2).strip()
            key = field_key(field_raw)
            # 可选值：第 3 列起（索引 2）的所有非空值，横向排列
            choices = [str(v).strip() for v in row[2:] if v is not None and str(v).strip()]
            valid_values.setdefault(key, []).append((field_raw, choices))
    return valid_values, groups


# --------------------------------------------------------------------------- #
# 解析主模板的 settings 行（row1）取出 labelRow/attributeRow/dataRow
# --------------------------------------------------------------------------- #
def parse_settings(settings_str: str) -> dict:
    params = {}
    for part in str(settings_str).split("&"):
        if "=" in part:
            k, _, v = part.partition("=")
            params[k] = v
    out = {}
    for key in ("labelRow", "attributeRow", "dataRow"):
        try:
            out[key] = int(params[key])
        except (KeyError, ValueError):
            out[key] = None
    return out


# --------------------------------------------------------------------------- #
# 解析主模板：按列收集 表头/属性键/示例数据
# --------------------------------------------------------------------------- #
def collect_column_data(ws, label_row, attr_row, data_row, max_samples=2):
    """
    读取主模板并按列组织成列表：
      [ (列号, 表头, 属性键, 示例值列表[…]), … ]
    只有表头非空且数据行（dataRow及之后）有值的列才进入结果。
    """
    rows = {}
    max_r = label_row
    # 遍历自 label_row 起的所有行，以便检测整列是否有数据
    for r, row in enumerate(
            ws.iter_rows(min_row=label_row, values_only=True),
            start=label_row,
    ):
        rows[r] = row
        max_r = r

    label_row_vals = rows.get(label_row, ())
    attr_row_vals = rows.get(attr_row, ())
    ncols = len(label_row_vals)
    columns = []
    for c in range(1, ncols + 1):
        label = label_row_vals[c - 1]
        if label is None or not str(label).strip():
            continue
        attr = attr_row_vals[c - 1] if c - 1 < len(attr_row_vals) else None

        # 扫描 dataRow 起的所有行，检查是否有值并提取示例数据
        samples = []
        has_data = False
        if data_row is not None:
            for r in range(data_row, max_r + 1):
                rr = rows.get(r, ())
                if c - 1 < len(rr) and rr[c - 1] is not None and str(rr[c - 1]).strip():
                    has_data = True
                    if len(samples) < max_samples:
                        samples.append(str(rr[c - 1]).strip())

        # 若数据行全为空，则不收集该列
        if not has_data:
            continue

        columns.append((c, str(label).strip(), str(attr).strip() if attr else "", samples))
    return columns


# --------------------------------------------------------------------------- #
# 默认输出文件名（自命名）：<输入文件名去扩展名>_analysis.json
# --------------------------------------------------------------------------- #
def default_output_path(source_file: str) -> str:
    """自命名默认输出：与输入文件同目录，文件名 = <输入文件名不含扩展名>_analysis.json"""
    base = os.path.splitext(os.path.basename(source_file))[0]
    directory = os.path.dirname(os.path.abspath(source_file))
    return os.path.join(directory, f"{base}_analysis.json")


# --------------------------------------------------------------------------- #
# 主流程
# --------------------------------------------------------------------------- #
def main():
    parser = argparse.ArgumentParser(
        description="只读解析 Amazon 模板 (.xlsm)，以 JSON 输出表头及字段可选值。"
    )
    parser.add_argument(
        "file",
        nargs="?",
        default=DEFAULT_INPUT_PATH,
        help="要解析的 .xlsm 文件路径（默认使用 DEFAULT_INPUT_PATH）",
    )
    parser.add_argument(
        "--output", "-o",
        default=DEFAULT_OUTPUT_PATH,
        help="JSON 输出文件路径（默认使用 DEFAULT_OUTPUT_PATH 或自命名）",
    )
    parser.add_argument(
        "--sheets",
        nargs="*",
        default=DEFAULT_SHEETS,
        help="参与解析的工作表名（默认按国家配置：fr=Valeurs valides/Modèle，de=Gültige Werte/Vorlage）",
    )
    parser.add_argument(
        "--max-values",
        type=int,
        default=0,
        help="每个字段最多输出的可选值个数，0 表示全部（默认）",
    )
    parser.add_argument(
        "--show-data",
        action="store_true",
        help="JSON 中附带每个表头列下的示例数据（samples 字段）",
    )
    args = parser.parse_args()

    # 日志统一写入 log/{当天日期}_atl_{用户}.log，不再打印控制台
    setup_log()

    try:
        wb = openpyxl.load_workbook(
            args.file, read_only=True, keep_vba=True, data_only=True
        )
    except Exception as exc:  # noqa: BLE001
        print(f"无法打开模板文件: {exc}")
        sys.exit(1)

    # 角色固定：字段枚举表(valid) 提供可选值，主模板(model) 提供表头，不依赖参数顺序
    sheets = zcfg.SHEET_NAMES[zcfg.COUNTRY]
    candidates = args.sheets or list(sheets.values())
    name_valid = next((s for s in candidates if s == sheets["valid"]), None)
    name_model = next((s for s in candidates if s == sheets["model"]), None)
    if name_valid is None or name_model is None:
        print(f"未找到工作表，请确认存在 {list(sheets.values())}（允许通过 --sheets 指定名称）。")
        sys.exit(2)

    # ---------- 1) 字段枚举表：字段 -> 可选值 ----------
    ws_valid = wb[name_valid]
    valid_values, groups = parse_valid_values(ws_valid)

    # ---------- 2) 主模板：定位表头/属性/数据行 ----------
    ws_model = wb[name_model]
    settings_row_vals = None
    for r, row in enumerate(ws_model.iter_rows(min_row=1, max_row=3, values_only=True), start=1):
        if row and row[0] is not None and str(row[0]).startswith("settings="):
            settings_row_vals = row
            break
    settings = {}
    if settings_row_vals and settings_row_vals[0]:
        settings = parse_settings(settings_row_vals[0])

    label_row = settings.get("labelRow") or 4
    attr_row = settings.get("attributeRow") or 5
    data_row = settings.get("dataRow") or 7

    label_row = max(1, int(label_row))
    attr_row = max(1, int(attr_row))
    data_row = int(data_row) if data_row and int(data_row) > 0 else (attr_row + 2)

    # ---------- 3) 汇总为 JSON 结构 ----------
    columns = collect_column_data(ws_model, label_row, attr_row, data_row)
    n_matched = 0
    columns_json = []
    used_keys: dict[str, int] = {}   # 同名表头列的出现次数 → 按顺序取同名候选字段
    for (col, label, attr, samples) in columns:
        key = normalize(label)
        entries = valid_values.get(key) or []
        k = used_keys.get(key, 0)          # 本列是同名表头中的第几个（0 起）
        used_keys[key] = k + 1
        hit = entries[k] if k < len(entries) else None
        # 每列预留两个标识位（自命名，不带数字），当前未使用，后续可填任意标记供程序复用
        reserve_flag = None
        reserve_mark = None
        if hit is None:
            col_entry = {
                "col": col,
                "header": label,
                "attribute": attr or None,
                "matched": False,
                "choices": [],
                "truncated": False,
                "samples": samples if args.show_data else [],
                "reserve_flag": reserve_flag,
                "reserve_mark": reserve_mark,
            }
        else:
            field_raw, choices = hit
            n_matched += 1
            truncated = False
            if args.max_values and args.max_values > 0 and len(choices) > args.max_values:
                choices = choices[: args.max_values]
                truncated = True
            col_entry = {
                "col": col,
                "header": label,
                "attribute": attr or None,
                "matched": True,
                "field": field_raw,
                "choices": choices,
                "truncated": truncated,
                "samples": samples if args.show_data else [],
                "reserve_flag": reserve_flag,
                "reserve_mark": reserve_mark,
            }
        columns_json.append(col_entry)

    result = {
        "source_file": args.file,
        "sheets": {"valid_values": name_valid, "model": name_model},
        "settings": settings,
        "groups": groups,
        "columns": columns_json,
        "stats": {"total": len(columns_json), "matched": n_matched},
    }

    out_path = args.output or default_output_path(args.file)
    os.makedirs(os.path.dirname(os.path.abspath(out_path)) or ".", exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"✅ 已输出: {out_path}")
    print(f"📊 表头 {len(columns_json)} 列，其中 {n_matched} 列在 {name_valid} 中找到可选值（未修改原文件）")


if __name__ == "__main__":
    main()