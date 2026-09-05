# -*- coding: utf-8 -*-
"""
ai_pick_attributes —— 对「有可选值的未覆盖列」用 **DeepSeek 网页版** 分批询问 AI 选值，
更新 M 数据源 JSON。

按 MISSING.md 的解决方案（类似 tools/title_optimize/deepseek_web.py 的网页自动化方式）：
  - 未覆盖列中，completed.json 有可选值(choices)的列
    → 提示词里含 产品列表 + 每列的可选值，AI 按列分块输出，每个产品每列选出 1 个值；
    每列的块标签直接用**原始表头**（归一化后），不转换语言；
  - **分批询问**：产品按每批 BATCH_SIZE 条（默认 10，--batch-size 可调，0 = 全部一批；
    最后一批按实际剩余条数）拆分，每批一个独立提示词/回答文件——避免一次问完所有产品
    导致回答超长、超时或截断；
  - 每列每组只存 1 个值 → fill_from_plan 以 cycle 模式循环铺满整组；
  - 无可选值的列保持 dataTemp 占位不变。

产物文件（ai_prompt/ 下，每批 2 个）：
  - attributes_batch01.txt / attributes_batch02.txt ...          各批提示词
  - attributes_batch01_result.txt / attributes_batch02_result.txt ...
    AI 回答（该批文件存在则复用该批，跳过网页）

只有网页版（无 API）：playwright 打开 DeepSeek 网页逐批发送提示词，读取回答解析写回。
登录态保存在 antelope/web_data/ + .deepseek_state.json（首次弹浏览器手动登录一次）。

用法:
    python antelope/ai_pick_attributes.py                          # 网页自动化 + 写回
    python antelope/ai_pick_attributes.py --batch-size 10          # 每批 10 条（默认）
    python antelope/ai_pick_attributes.py --products-file products.txt   # 用指定产品列表
    python antelope/ai_pick_attributes.py --generate-only          # 只生成各批提示词文件，不发网页
    python antelope/ai_pick_attributes.py --source-excel x --m-data y
"""

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

import openpyxl

from common import (
    load_ai_columns,
    load_data_cols,
    load_groups,
    load_json,
    setup_log,
    uncovered_cols,
    zcfg,
)

BASE = Path(__file__).resolve().parent

DEFAULT_DIFF = zcfg.CFG_INTERMEDIATE["column_diff_json"]
DEFAULT_COMPLETED = zcfg.CFG_INTERMEDIATE["completed_json"]
DEFAULT_GROUPS = zcfg.CFG_INTERMEDIATE["groups_json"]
DEFAULT_DATA = zcfg.CFG_INTERMEDIATE["data_json"]
DEFAULT_SOURCE_EXCEL = zcfg.DATA_SOURCE_A
DEFAULT_M_DATA = zcfg.DATA_SOURCE_M
DEFAULT_PROMPT_DIR = zcfg.CFG_INTERMEDIATE["ai_prompt_dir"]

# 产品描述取数据源 A(Sheet0) 的这些列：I(9)=中文标题（产品列表展示用中文）
PRODUCT_DESC_COLS = (9,)

# ── DeepSeek 网页自动化 ──
DEEPSEEK_URL = "https://chat.deepseek.com/"
WEB_DATA_DIR = BASE / "web_data"            # 浏览器数据（登录态持久化）
STATE_FILE = BASE / ".deepseek_state.json"  # 登录态快照（复用）
ASSISTANT_SEL = "div.ds-assistant-message-main-content"

# ── 分批询问 ──
# 每批产品条数（最后一批按实际剩余）；0 = 全部一批（旧行为）
BATCH_SIZE = 10


def batch_file_names(idx: int) -> tuple[str, str]:
    """第 idx 批（1 起）的 提示词文件名 / 回答文件名。

    例: idx=1 -> ("attributes_batch01.txt", "attributes_batch01_result.txt")
    """
    tag = f"batch{idx:02d}"
    return f"attributes_{tag}.txt", f"attributes_{tag}_result.txt"


def build_batches(products, batch_size: int) -> list[tuple[int, list]]:
    """把产品列表按每批 batch_size 条拆分为 [(起始序号(0 起), 本批产品列表), ...]。

    最后一批按实际剩余条数；batch_size <= 0 时视为全部一批。
    """
    if batch_size is None or batch_size <= 0:
        batch_size = max(len(products), 1)
    return [
        (start, products[start : start + batch_size])
        for start in range(0, len(products), batch_size)
    ]


def normalize_label(text: str, fold_case: bool = False) -> str:
    """归一化标签文本，用于块头匹配与提示词展示。

    模板表头里常带 \\xa0（不换行空格）、\\u2007、\\u202f 以及弯引号/全角标点，
    AI 照抄时容易转成普通空格/半角，导致精确匹配失败、整块串列（如
    'Artikeltyp\\xa0– Name' vs 'Artikeltyp – Name'）。这里统一：
      1. 所有不可见空白（普通空格、\\xa0、\\u2007、\\u202f）→ 普通空格，并压缩、去首尾空白；
      2. 弯引号/零宽空格 → ASCII 引号/删除；
      3. fold_case=True 时再 casefold（解析端匹配用，容忍 AI 大小写差异）；
      4. 生成提示词时用 fold_case=False，保留原始大小写便于 AI 照抄。
    """
    s = str(text)
    s = re.sub(r"[\s\u00a0\u2007\u202f]+", " ", s).strip()
    for a, b in (("\u2018", "'"), ("\u2019", "'"), ("\u201a", "'"),
                 ("\u201c", '"'), ("\u201d", '"'), ("\u201e", '"'),
                 ("\u200b", "")):
        s = s.replace(a, b)
    if fold_case:
        s = s.casefold()
    return s.strip()


def make_labels(ai_cols) -> dict[int, str]:
    """给每列分配标签：直接用原始表头（归一化后的原文），不转换；同名表头自动加序号（如 Style / Style2）。"""
    used: dict[str, int] = {}
    labels: dict[int, str] = {}
    for col, header, _ in ai_cols:
        base = normalize_label(header)
        n = used.get(base, 0) + 1
        used[base] = n
        labels[col] = f"{base}{n}" if n > 1 else base
    return labels


def build_product_list_from_excel(source_excel, groups):
    """从数据源 A(Sheet0) 提取产品列表：每组取锚点行（组起始行）的 标题|卖点。"""
    wb = openpyxl.load_workbook(source_excel, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]  # 规则：只读第一个工作表（Sheet0）
        products = []
        for gname, spec in (groups or {}).items():
            spec = str(spec).strip()
            if "&" not in spec:
                continue
            try:
                start_actual, _ = map(int, spec.split("&"))
            except ValueError:
                continue
            parts = []
            for c in PRODUCT_DESC_COLS:
                v = ws.cell(row=start_actual, column=c).value
                if v is not None and str(v).strip():
                    parts.append(str(v).strip())
            products.append(" | ".join(parts) if parts else f"({gname})")
        return products
    finally:
        wb.close()


def load_products_from_file(path):
    """从文本文件读取产品列表（每行一个产品）。"""
    return [l.strip() for l in Path(path).read_text(encoding="utf-8").splitlines() if l.strip()]


def build_prompt_all(products, ai_cols, labels):
    """构造一次合并的提示词：产品列表 + 各组可选值（编号形式，AI 输出编号杜绝造词）。"""
    lines = [
        "请为以下每个产品，从各组可选值中选出最合适的 1 个值，必须选。",
        "",
        "【规则】",
        "1. 每个值都必须是该组「可选值」列表中的原词；若没有完全匹配的值，请在**该组列表内**选择语义上最接近、最合适的一个或者选择通用的泛化的。必须在组内选一个",
        "2. 为方便你作答，可选值已编号；**输出时直接给出选项编号（数字）**，不要输出文字值，不要造词。",
        "",
        "【产品列表】（每组一个产品，编号 1~N）：",
    ]
    for i, p in enumerate(products, 1):
        lines.append(f"{i}. {p}")
    lines += [
        "",
        "【可选值组】（每组只能从该组编号中选择）：",
    ]
    for col, header, choices in ai_cols:
        numbered = ", ".join(f"{i}.{c}" for i, c in enumerate(choices, 1))
        lines.append(f"{labels[col]}: {numbered}")
    lines += [
        "",
        "【输出格式】严格按下面列出的块头顺序输出：每个块头占一行（以冒号结尾，**逐字照抄块头，不要改写、不要加序号**），",
        "块内每行一个产品（产品编号+冒号+选项编号），不要任何解释、不要增删块头：",
    ]
    for c, _, _ in ai_cols:
        lines.append(f"{labels[c]}:")
    lines += ["1: <选项编号>", "2: <选项编号>", "..."]
    return "\n".join(lines)


_LINE_RE = re.compile(r"^\s*(\d+)\s*[:：]\s*(.+?)\s*$")


def parse_all(text, ai_cols, n_products, labels):
    """解析一次合并回答 → {col: {产品编号: 值}}。

    值支持两种形态：
      - 数字（推荐，提示词要求输出选项编号）→ 映射回该列可选值；
      - 文字 → 严格匹配该列可选值（大小写/变体归一化）。
    非法值（编号越界 / 文字不在可选值内）→ 拒绝并警告，该产品保持占位。
    """
    choices_by_col = {col: [str(c) for c in choices] for col, _, choices in ai_cols}
    norm = {col: {str(c).strip().casefold(): str(c).strip() for c in choices}
            for col, _, choices in ai_cols}
    # 块头匹配：两侧都做归一化（\xa0/空白/引号/大小写），避免 AI 转写差异导致整块串列
    label_to_col = {normalize_label(labels[col], fold_case=True): col for col, _, _ in ai_cols}
    result: dict[int, dict[int, str]] = {}
    rejected: list[tuple[int, str]] = []   # (col, 被拒值)

    cur_col = None
    unknown_blocks: list[str] = []   # 回答中出现但未匹配任何列标签的块头（疑似改写/丢列）
    for raw in str(text or "").splitlines():
        raw = raw.strip()
        if not raw:
            continue
        line = raw.rstrip(":：").strip()
        if not line:
            continue
        key = normalize_label(line, fold_case=True)
        if key in label_to_col:          # 块头（标签，归一化后匹配）
            cur_col = label_to_col[key]
            continue
        ml = _LINE_RE.match(line)
        if ml is None:
            # 非值行：若以冒号结尾且形似标签（含拉丁字母、短文本）→ 未识别块头，整块会丢，必须告警
            if raw.endswith(":") or raw.endswith("："):
                prefix = raw.rstrip(":：").strip()
                if prefix and len(prefix) <= 50 and re.search(r"[A-Za-z\u00c0-\u024f]", prefix):
                    unknown_blocks.append(raw)
            continue
        if cur_col is None:
            continue
        no = int(ml.group(1))
        val = ml.group(2).strip()
        if not (1 <= no <= n_products and val):
            continue

        # 形态 1：选项编号
        if val.isdigit():
            idx = int(val)
            opts = choices_by_col.get(cur_col, [])
            if 1 <= idx <= len(opts):
                result.setdefault(cur_col, {})[no] = opts[idx - 1]
            else:
                rejected.append((cur_col, val))
            continue

        # 形态 2：文字，严格匹配可选值
        matched = norm.get(cur_col, {}).get(val.casefold())
        if matched is None:
            rejected.append((cur_col, val))
            continue
        result.setdefault(cur_col, {})[no] = matched

    if unknown_blocks:
        print(f"  ⚠️ 回答中出现 {len(unknown_blocks)} 个未识别块头（未匹配任何列，整块将丢失）: {unknown_blocks}")
    if rejected:
        print(f"  ⚠️ 拒绝 {len(rejected)} 个不在可选值内的值（保持占位）: {rejected}")
    return result


def update_m_data(m_data, groups, col, parsed, placeholder, covered=None):
    """把 AI 选值写入 M 数据（每组 1 个值 → cycle 铺满）。

    covered: 本批覆盖的绝对组序号集合（1 起）。分批调用时**只处理这些组**，
    其余组保持原值不动——否则后跑的批次会把前面批次已写入的真实值覆盖回占位。
    covered 为 None 时（单批/旧行为）处理全部组。

    本批覆盖但无合法值（AI 缺失/被拒）→ **覆盖为占位**（不留旧值残留）。
    """
    updated = 0
    kept = 0
    for i, (gname, spec) in enumerate((groups or {}).items(), 1):
        if str(spec).strip().count("&") != 1:
            continue
        if covered is not None and i not in covered:
            continue                       # 本批未覆盖的组：保持原值，不动
        gdata = m_data.setdefault(str(gname), {})
        val = parsed.get(i)
        if val:
            gdata[str(col)] = [val]          # 单元素 → fill_from_plan cycle 铺满
            updated += 1
        else:
            gdata[str(col)] = [placeholder]  # 覆盖为占位，防止旧值残留
            kept += 1
    return updated, kept


# ══════════════════════════════════════════════════════════════════════════
#  DeepSeek 网页自动化（参考 tools/title_optimize/deepseek_web.py）
# ══════════════════════════════════════════════════════════════════════════
def last_assistant_text(page) -> str:
    """页面上最后一条助手回复文本（发送前记录基线用）。"""
    locator = page.locator(ASSISTANT_SEL)
    texts = []
    for m in locator.all():
        try:
            t = m.inner_text().strip()
        except Exception:
            continue
        if t:
            texts.append(t)
    return texts[-1] if texts else ""


def extract_last_response(page, previous_text: str = "", timeout: int = 300) -> str:
    """等待生成结束并提取最后一条助手回复（排除思考过程；新文本连续 3 秒无变化判定完成）。"""
    locator = page.locator(ASSISTANT_SEL)
    stop_btn = page.locator("button:has-text('Stop')")
    deadline = time.monotonic() + timeout
    last_text, stable = "", 0
    while time.monotonic() < deadline:
        texts = []
        for m in locator.all():
            try:
                t = m.inner_text().strip()
            except Exception:
                continue
            if t:
                texts.append(t)
        cur = texts[-1] if texts else ""
        if cur and cur == last_text:
            stable += 1
            if stable >= 3 and cur != previous_text and not stop_btn.is_visible():
                return cur
        else:
            stable = 0
        last_text = cur
        time.sleep(1)
    if last_text and last_text != previous_text:
        return last_text
    raise RuntimeError("No new assistant response found")


def validate_no_placeholder(m_data, ai_cols, labels, placeholder):
    """校验：有可选值的列不允许残留占位（dataTemp）。

    任一 AI 列仍有组是占位（本批回答缺失/块头未识别等）→ 逐列列出残留组数，
    便于补齐对应批次回答后重跑。
    """
    bad = {}
    for col, header, _ in ai_cols:
        n_ph = 0
        for gdata in m_data.values():
            v = (gdata or {}).get(str(col))
            if v is None:
                n_ph += 1
            elif all(str(x) == placeholder for x in v):
                n_ph += 1
        if n_ph:
            bad[col] = n_ph
    if bad:
        print("⚠️ 仍有可选值列残留占位（请补齐对应批次回答后重跑本程序）:")
        for col, n in bad.items():
            print(f"   列{col} ({labels.get(col, '')}): {n} 组仍为占位")
    else:
        print("✅ 校验通过：所有有可选值列均无占位")


def web_ask_all(ai_cols, products, args, m_data, groups):
    """用 DeepSeek 网页分批询问所有产品（每批 BATCH_SIZE 条），解析后更新 M JSON。

    每批一个提示词/回答文件；该批回答文件已存在则直接复用，跳过网页。
    需要网页询问的批次共用一次浏览器会话，逐批发送。
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("❌ playwright 未安装。请先执行：")
        print("    pip install playwright")
        print("    playwright install chromium")
        sys.exit(3)

    os.makedirs(args.prompt_dir, exist_ok=True)
    WEB_DATA_DIR.mkdir(parents=True, exist_ok=True)

    labels = make_labels(ai_cols)
    batches = build_batches(products, args.batch_size)
    print(f"产品共 {len(products)} 个，按每批 {args.batch_size if args.batch_size and args.batch_size > 0 else '全部'} 个"
          f"分为 {len(batches)} 批（最后一批 {len(batches[-1][1])} 个）")

    # 先写出各批提示词并检查回答文件：已存在的批直接复用，其余记入 todo 待网页询问
    todo = []
    for b, (start, batch) in enumerate(batches, 1):
        p_name, r_name = batch_file_names(b)
        p_path = os.path.join(args.prompt_dir, p_name)
        r_path = os.path.join(args.prompt_dir, r_name)
        prompt = build_prompt_all(batch, ai_cols, labels)
        Path(p_path).write_text(prompt, encoding="utf-8")
        print(f"📄 提示词: {p_name}（产品 {start + 1}~{start + len(batch)}）")
        covered = set(range(start + 1, start + len(batch) + 1))  # 本批覆盖的绝对组序号

        if os.path.exists(r_path):
            text = Path(r_path).read_text(encoding="utf-8")
            parsed = parse_all(text, ai_cols, len(batch), labels)
            total_u = total_k = 0
            for col, _, _ in ai_cols:
                abs_parsed = {start + j: v for j, v in parsed.get(col, {}).items()}
                u, k = update_m_data(m_data, groups, col, abs_parsed, args.placeholder, covered=covered)
                total_u += u
                total_k += k
            missing = [f"{c}({labels[c]})" for c, _, _ in ai_cols if not parsed.get(c)]
            if missing:
                print(f"  ⚠️ 本批列 {missing} 未解析到任何值，将保持占位（请检查对应块头）")
            print(f"✅ 批次 {b}/{len(batches)}（复用 {r_name}）：写入 {total_u} 个「组×列」，占位 {total_k}")
        else:
            todo.append((b, start, batch, p_path, r_path, prompt))

    if not todo:
        print("所有批次均已有回答结果，无需打开网页。")
        validate_no_placeholder(m_data, ai_cols, labels, args.placeholder)
        return

    # 需要网页询问的批次：一次性打开浏览器，逐批发送
    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=str(WEB_DATA_DIR),
            headless=False,
        )
        page = context.new_page()
        page.goto(DEEPSEEK_URL, wait_until="domcontentloaded")
        print(f"已打开 {DEEPSEEK_URL}")

        if not STATE_FILE.exists():
            print("=" * 60)
            print("首次运行：请在弹出的浏览器中登录 DeepSeek")
            print("登录完成后程序会自动继续（最多等待 10 分钟）...")
            print("=" * 60)
            logged_in = False
            for _ in range(600):
                try:
                    if page.locator("textarea, [contenteditable]").count() > 0:
                        logged_in = True
                        break
                except Exception:
                    pass
                time.sleep(1)
            if not logged_in:
                raise RuntimeError("等待登录超时（10 分钟），请确认浏览器中已登录 DeepSeek")
            time.sleep(2)
            context.storage_state(path=str(STATE_FILE))
            print(f"登录态已保存: {STATE_FILE.name}")
        else:
            print(f"复用登录态: {STATE_FILE.name}")

        for b, start, batch, p_path, r_path, prompt in todo:
            covered = set(range(start + 1, start + len(batch) + 1))  # 本批覆盖的绝对组序号
            print(f"--- 批次 {b}/{len(batches)}：产品 {start + 1}~{start + len(batch)}（共 {len(batch)} 个）---")
            page.wait_for_selector("textarea, [contenteditable]", timeout=30000)
            time.sleep(2)
            print("DeepSeek 对话页就绪")

            new_chat_btn = page.locator("text=New Chat").first
            if new_chat_btn.is_visible():
                new_chat_btn.click()
                time.sleep(1)

            textarea = page.locator("textarea").first
            textarea.fill(prompt)
            time.sleep(1)

            prev_text = last_assistant_text(page)
            page.keyboard.press("Enter")
            print("已发送，等待 AI 回答...")
            time.sleep(15)

            answer = extract_last_response(page, previous_text=prev_text)
            Path(r_path).write_text(answer, encoding="utf-8")
            print(f"📄 回答已保存: {r_name}")

            parsed = parse_all(answer, ai_cols, len(batch), labels)
            total_u = total_k = 0
            for col, _, _ in ai_cols:
                abs_parsed = {start + j: v for j, v in parsed.get(col, {}).items()}
                u, k = update_m_data(m_data, groups, col, abs_parsed, args.placeholder, covered=covered)
                total_u += u
                total_k += k
            missing = [f"{c}({labels[c]})" for c, _, _ in ai_cols if not parsed.get(c)]
            if missing:
                print(f"  ⚠️ 本批列 {missing} 未解析到任何值，将保持占位（请检查对应块头）")
            print(f"✅ 批次 {b}/{len(batches)}（网页询问）：写入 {total_u} 个「组×列」，占位 {total_k}")

        context.close()

    validate_no_placeholder(m_data, ai_cols, labels, args.placeholder)


def main():
    parser = argparse.ArgumentParser(
        description="对「有可选值的未覆盖列」用 DeepSeek 网页版一次询问 AI 选值，更新 M 数据源 JSON"
    )
    parser.add_argument("--diff", default=DEFAULT_DIFF, help="column_diff JSON 路径")
    parser.add_argument("--completed", default=DEFAULT_COMPLETED, help="completed 分析 JSON 路径")
    parser.add_argument("--groups", default=DEFAULT_GROUPS, help="groups JSON 路径")
    parser.add_argument("--data", default=DEFAULT_DATA, help="A 取数 data JSON 路径（判断已覆盖列）")
    parser.add_argument("--source-excel", default=DEFAULT_SOURCE_EXCEL, help="数据源 A(.xlsx)，无 --products-file 时用它提取产品列表")
    parser.add_argument("--products-file", default=None, help="产品列表文件（每行一个产品）；缺省从数据源 A 提取")
    parser.add_argument("--m-data", default=DEFAULT_M_DATA, help="M JSON 路径（读占位/写结果）")
    parser.add_argument("--prompt-dir", default=DEFAULT_PROMPT_DIR,
                        help="提示词/结果文件目录（每批 2 个文件：attributes_batchNN.txt + _result.txt）")
    parser.add_argument("--placeholder", default="dataTemp", help="占位值")
    parser.add_argument("--batch-size", type=int, default=BATCH_SIZE,
                        help="每批询问的产品条数（默认 10；0 = 全部一批；最后一批按实际剩余）")
    parser.add_argument("--generate-only", action="store_true",
                        help="只生成各批提示词文件，不打开网页（手动喂 AI，回答存对应 _result.txt 后重跑写回）")
    args = parser.parse_args()

    setup_log()

    groups = load_groups(args.groups)
    if not groups:
        print("⚠️ groups 为空，无法确定分组行数")
        sys.exit(1)

    ai_cols = load_ai_columns(args.completed, uncovered_cols(args.diff, args.data))

    # 识别核对：completed 里有可选值、但被 A 覆盖（不进 AI，由 A 数据填充）的列，列出便于人工确认
    try:
        comp_cols = {c["col"]: c for c in load_json(args.completed).get("columns", [])}
        a_cols = load_data_cols(args.data)
        skipped = [
            (c["col"], c.get("header"))
            for c in comp_cols.values()
            if (c.get("choices") or []) and str(c["col"]) in a_cols
        ]
        if skipped:
            print(f"ℹ️ 有可选值但由 A 覆盖、不进 AI 的列（由 A 数据填充）: {skipped}")
    except Exception:
        pass  # 诊断信息失败不影响主流程

    if not ai_cols:
        print("未覆盖列中没有找到有可选值(choices)的列，无需 AI 选值。")
        return

    # 产品列表
    if args.products_file:
        products = load_products_from_file(args.products_file)
        print(f"产品列表（文件）：{len(products)} 个")
    else:
        products = build_product_list_from_excel(args.source_excel, groups)
        print(f"产品列表（数据源 A Sheet0 提取）：{len(products)} 个")

    print("🤖 AI 选值列（有可选值的未覆盖列，按批询问，每批 "
          f"{args.batch_size if args.batch_size and args.batch_size > 0 else '全部'} 条）:")
    for col, header, choices in ai_cols:
        print(f"   {col}: {header}（{len(choices)} 个可选值）")

    # 读现有 M JSON
    m_data = {}
    if os.path.exists(args.m_data):
        m_data = load_json(args.m_data).get("data") or {}

    if args.generate_only:
        os.makedirs(args.prompt_dir, exist_ok=True)
        labels = make_labels(ai_cols)
        batches = build_batches(products, args.batch_size)
        for b, (start, batch) in enumerate(batches, 1):
            p_name, _ = batch_file_names(b)
            p_path = os.path.join(args.prompt_dir, p_name)
            Path(p_path).write_text(build_prompt_all(batch, ai_cols, labels), encoding="utf-8")
            print(f"📄 提示词: {p_name}（产品 {start + 1}~{start + len(batch)}）")
        print(f"✅ 共 {len(batches)} 批提示词已生成；把各批喂给 AI，回答存为对应 _result.txt 后重跑本程序写回。")
        return

    web_ask_all(ai_cols, products, args, m_data, groups)

    # 写回 M JSON
    with open(args.m_data, "w", encoding="utf-8") as f:
        json.dump({"data": m_data, "placeholder": args.placeholder}, f, ensure_ascii=False, indent=2)
    print(f"\nM 数据源已更新: {args.m_data}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"错误: {e}", file=sys.stderr)
        sys.exit(1)
