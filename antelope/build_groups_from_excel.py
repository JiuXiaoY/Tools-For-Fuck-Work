# -*- coding: utf-8 -*-
"""
从「数据源 A(.xlsx)」生成 groups（填充计划所需的行范围）。

按 11409 需求的分工（A/B/C/M/D 角色见 zconfig.constant.py）：
  输入 = 数据源 A（.xlsx_dataSource），第一列（A 列）有可见背景填充色的单元格 = 父体锚点行。

分组**没有偏移**：groups 里直接存 excel 的实际行号（锚点行原样，不做相对化换算）；
偏移（data_start_row）只在「填充数据」时由 fill_from_plan.py 应用。

约定：
  - 读取 excel 第一列（A 列），有可见背景填充色的单元格所在行 = 父体锚点行；
  - 每个锚点行 = 一个组的起始行，组结束行 = 下一个锚点行 - 1；
  - 最后一个组的结束行 = excel 最大行（max_row）；
  - groups 形如 { "group_1": "1 & 19", "group_2": "20 & 26" }（实际行号，含起始行）。

有色判断复用 services/utils.py 的 cell_has_fill（排除白色/黑色/透明等非可见填充）。

用法:
    python build_groups_from_excel.py [excel 路径] [-o output.json]
        [--scan-from N] [--end-row N]
默认:
    input  = xlsm/.xlsx_dataSource（数据源 A）
    output = intermediate/fr_shirt/fr_shirt_groups.json（自命名）
"""
import argparse
import json
import os
import sys
from pathlib import Path

import openpyxl

from common import cell_has_fill, setup_utf8, zcfg

BASE_DIR = Path(__file__).resolve().parent
INTERMEDIATE_DIR = Path(zcfg.INTERMEDIATE_DIR)

DEFAULT_INPUT_PATH = zcfg.DATA_SOURCE_A   # 数据源 A（.xlsx_dataSource）
DEFAULT_OUTPUT_PATH = Path(zcfg.CFG_INTERMEDIATE["groups_json"])


def resolve_input(path: str) -> Path:
    """输入可以是具体 .xlsx 文件，也可以是目录（取其中最新的 .xlsx）。"""
    p = Path(path)
    if p.is_dir():
        candidates = sorted(p.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
        if not candidates:
            raise FileNotFoundError(f"目录 {p} 下没有 .xlsx 文件")
        return candidates[0]
    if not p.exists():
        raise FileNotFoundError(f"找不到 excel 文件: {p}")
    return p


def find_anchor_rows(ws, scan_from: int) -> list[int]:
    """扫描第一列，返回所有有色单元格的行号（升序）。

    注意：必须在 read_only 模式下用 iter_rows 顺序遍历；
    用 ws.cell() 随机访问会导致大文件性能灾难。
    """
    anchors = []
    for row in ws.iter_rows(min_col=1, max_col=1):
        r = row[0].row
        if r < scan_from:
            continue
        if cell_has_fill(row[0]):
            anchors.append(r)
    return anchors


def build_groups(anchors: list[int], end_row: int) -> tuple[dict, list[dict]]:
    """由锚点行生成 groups（实际行号）及核对用的 details。

    分组无偏移：起始/结束行直接用 excel 实际行号。
    """
    groups = {}
    details = []
    for i, anchor in enumerate(anchors):
        gname = f"group_{i + 1}"
        group_end = (anchors[i + 1] - 1) if i + 1 < len(anchors) else end_row
        if group_end < anchor:
            group_end = anchor
        groups[gname] = f"{anchor} & {group_end}"
        details.append({
            "group": gname,
            "anchor_row": anchor,
            "actual_rows": f"{anchor} & {group_end}",
            "row_count": group_end - anchor + 1,
        })
    return groups, details


def main():
    parser = argparse.ArgumentParser(description="从数据源 A(.xlsx) 的第一列有色单元格生成 groups")
    parser.add_argument("input", nargs="?", default=DEFAULT_INPUT_PATH,
                        help="数据源 A(.xlsx) 路径（默认 xlsm/.xlsx_dataSource）")
    parser.add_argument("-o", "--output", default=str(DEFAULT_OUTPUT_PATH),
                        help="输出 JSON 文件路径")
    parser.add_argument("--scan-from", type=int, default=1,
                        help="从第几行开始扫描第一列有色单元格（默认 1）")
    parser.add_argument("--end-row", type=int, default=0,
                        help="最后一个组的结束行（默认取 excel 的 max_row）")
    parser.add_argument("--verbose", action="store_true",
                        help="逐组打印分组行范围明细")
    args = parser.parse_args()

    setup_utf8()

    src = resolve_input(args.input)
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    try:
        # 规则：.xlsx 数据源只读第一个工作表（Sheet0），其余忽略
        ws = wb[wb.sheetnames[0]]
        end_row = args.end_row if args.end_row > 0 else ws.max_row
        anchors = find_anchor_rows(ws, args.scan_from)
        if not anchors:
            print(f"⚠️ 第一列未找到任何有色单元格（{src.name}）")
        groups, details = build_groups(anchors, end_row)
    finally:
        wb.close()

    result = {
        "groups": groups,
    }

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    total_rows = sum(d["row_count"] for d in details)
    print(f"✅ 分组完成：{len(anchors)} 个锚点 -> {len(groups)} 组，共 {total_rows} 行（实际行号，无偏移）")
    if args.verbose:
        for d in details:
            print(f"   {d['group']}: 行 {d['actual_rows']}（{d['row_count']} 行）")
    print(f"📄 已写入: {out_path}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"错误: {e}", file=sys.stderr)
        sys.exit(1)
