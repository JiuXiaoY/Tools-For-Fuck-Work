# -*- coding: utf-8 -*-
"""
从「数据源 A(.xlsx)」读取列数据，按列映射填充到 plan 的 data（按分组组织）。

按 11409 需求的分工（A/B/C/M/D 角色见 zconfig.constant.py）：
  - A 提供「部分」待填列数据：经 col_mapping（映射文件）取数；
  - 剩余未映射到的待填列由 M（自定义数据来源，JSON）补充 ——
    合并动作在 build_fill_framework.py 中完成（A 已有的列优先，M 只补空缺列）；
  - 分组行范围来自 groups.json（build_groups_from_excel.py 对 A 生成）。

流程与 build_groups_from_excel.py 对称：
  - groups 来源 json（默认 intermediate/fr_shirt/fr_shirt_groups.json）提供分组行范围（**实际行号，无偏移**）；
  - 列映射来源 json（默认 intermediate/fr_shirt/fr_shirt_col_mapping.json）提供「源 excel 列 → plan 目标列」映射；
  - 本程序读取数据源 excel：对每个分组、每个源列取「该组行范围内该列的非空值」，
    写入目标列；一对多映射（如 K→BY,BZ）表示同一份数据复制到多个目标列；
    映射应用到全部分组；
  - 生成 data 来源 JSON（默认 intermediate/fr_shirt/fr_shirt_data.json）：
        { "data": { <group>: { "<目标列号>": [ 值序列 ] } } }
    交给 build_fill_framework.py 读取（与 M 数据合并）后写入 plan 的 data 字段，
    再交给 fill_from_plan.py。

取数约定：
  - groups 里的行范围 = 数据源 A 的实际行号，直接按此读取，不做偏移；
  - 每个源列在该组行范围内**逐行读取，空单元格保留为 "" 占位**（m 含空数据）；
    未使用到的 A 列（不在映射里）忽略不读；映射到但没数据的列也读，空就是空；
  - 性能：只读模式先把所需行×列一次性顺序读入内存（snapshot_rows），再逐组取值，
    避免逐格 ws.cell() 随机访问（openpyxl 只读模式下每次都会从头重新解析整个 XML，
    行数一大（如 1500 行）会卡死数十分钟）。

用法:
    python build_data_from_excel.py [-o output.json]
        [--groups groups.json] [--col-mapping mapping.json]
        [--source-excel data.xlsx] [--sheet 工作表名]

默认:
    groups       = intermediate/fr_shirt/fr_shirt_groups.json
    col-mapping  = intermediate/fr_shirt/fr_shirt_col_mapping.json
    source-excel = xlsm/.xlsx_dataSource（数据源 A）
    output       = intermediate/fr_shirt/fr_shirt_data.json
"""
import argparse
import json
import os
import sys

import openpyxl
from openpyxl.utils import column_index_from_string

from common import load_groups, load_json, zcfg

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INTERMEDIATE_DIR = zcfg.INTERMEDIATE_DIR

DEFAULT_GROUPS = zcfg.CFG_INTERMEDIATE["groups_json"]
DEFAULT_COL_MAPPING = zcfg.CFG_INTERMEDIATE["col_mapping_json"]
DEFAULT_SOURCE_EXCEL = zcfg.DATA_SOURCE_A   # 数据源 A（.xlsx_dataSource）
DEFAULT_OUTPUT = zcfg.CFG_INTERMEDIATE["data_json"]


def load_col_mapping(path):
    """读取列映射来源 JSON 的 sources 数组。

    返回: [ {"name": str, "mapping": {源字母: [target字母,...]}} ]
    """
    if not path or not os.path.exists(path):
        return []
    try:
        return list(load_json(path).get("sources") or [])
    except Exception:
        return []


def parse_col_mapping(sources):
    """字母映射 -> 源列号到目标列号列表的字典。

    返回: { src_col(int): [target_col(int), ...] }
    """
    mapping = {}
    for src in sources:
        for src_letter, target_letters in (src.get("mapping") or {}).items():
            src_col = column_index_from_string(src_letter)
            targets = [column_index_from_string(t) for t in target_letters]
            mapping.setdefault(src_col, [])
            for t in targets:
                if t not in mapping[src_col]:
                    mapping[src_col].append(t)
    return mapping


def snapshot_rows(ws, max_row: int, max_col: int) -> dict:
    """一次性顺序读入所需行×列到内存，供后续取值（避免只读模式逐格随机访问的性能灾难）。

    openpyxl 只读模式下每次 ws.cell(row, col) 都会**从头重新解析整个工作表 XML**
    （见 ReadOnlyWorksheet._get_cell → _cells_by_row），行数一多（如 1570 行 × 18 个源列
    ≈ 2.8 万次随机访问）耗时会暴涨到几十分钟；改为一次 iter_rows 顺序流式读取（约 0.2s），
    之后全部走内存查找。

    返回 { 行号: 值元组 }，行号从 1 起；行内缺失的列返回 None（与 ws.cell() 的空单元格语义一致）。
    """
    rows = {}
    for r, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True),
        start=1,
    ):
        rows[r] = row
    return rows


def read_column_values(rows, group_start: int, group_end: int, src_col: int) -> list:
    """从内存行矩阵中读某组行范围内某列的值序列（**逐行读取，空单元格保留为 "" 占位**）。

    规则（MISSING.md 第 2 节）：映射到但没数据的列也要读取——空就是空，写入也是空。
    返回列表长度 = 组行数（m 含空数据，与 11409 需求一致）；空单元格记 ""。
    """
    values = []
    for r in range(group_start, group_end + 1):
        rr = rows.get(r, ())
        v = rr[src_col - 1] if 0 < src_col <= len(rr) else None
        if v is None:
            values.append("")
        elif isinstance(v, str) and not v.strip():
            values.append("")
        else:
            values.append(v)
    return values


def extract_data(wb, groups, col_mapping, sheet=None):
    """从工作簿按分组 + 列映射提取 data；映射应用于全部组。

    分组行范围 = 数据源 A 的实际行号，直接读取（无偏移）。
    返回: { group: { str(target_col): [值...] } }

    性能：只读模式先 snapshot_rows 一次性把所需范围读入内存，再逐组逐列取值，
    避免逐格 ws.cell() 随机访问（每次从头解析整个 XML，行数大时会卡死）。
    """
    if sheet and sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        # 规则：.xlsx 数据源只读第一个工作表（Sheet0），其余忽略
        ws = wb[wb.sheetnames[0]]

    # 计算所需行/列范围，一次性顺序读入内存
    max_row = 0
    for spec in (groups or {}).values():
        spec = str(spec).strip()
        if "&" not in spec:
            continue
        try:
            end = int(spec.split("&")[1])
        except ValueError:
            continue
        max_row = max(max_row, end)
    max_col = max(col_mapping) if col_mapping else 0
    rows = snapshot_rows(ws, max_row, max_col) if max_row > 0 else {}

    data = {}
    for gname, spec in (groups or {}).items():
        spec = str(spec).strip()
        if "&" not in spec:
            continue
        try:
            start_actual, end_actual = map(int, spec.split("&"))
        except ValueError:
            continue
        if start_actual < 1 or end_actual < start_actual:
            continue

        gdata = {}
        for src_col, target_cols in col_mapping.items():
            values = read_column_values(rows, start_actual, end_actual, src_col)
            for target in target_cols:
                gdata[str(target)] = list(values)  # 同一份数据复制到多个目标列
        data[gname] = gdata
        print(f"  [取数] {gname}: 行 {start_actual}..{end_actual} -> {len(gdata)} 个目标列")
    return data


def main():
    parser = argparse.ArgumentParser(
        description="从数据源 excel 按分组+列映射生成 data 来源 JSON"
    )
    parser.add_argument("-o", "--output", default=DEFAULT_OUTPUT,
                        help="输出 data 来源 JSON 路径")
    parser.add_argument("--groups", default=DEFAULT_GROUPS,
                        help="分组来源 JSON 路径（取 groups 字段，实际行号，无偏移）")
    parser.add_argument("--col-mapping", default=DEFAULT_COL_MAPPING,
                        help="列映射 JSON 路径（取 sources 数组中的字母映射）")
    parser.add_argument("--source-excel", default=DEFAULT_SOURCE_EXCEL,
                        help="数据源 excel 路径（默认 xlsm/.xlsx_dataSource，即数据源 A）")
    parser.add_argument("--sheet", default=None,
                        help="工作表名（默认只读第一个工作表 Sheet0；其他工作表忽略）")
    args = parser.parse_args()

    groups = load_groups(args.groups)
    sources = load_col_mapping(args.col_mapping)
    col_mapping = parse_col_mapping(sources)

    if not groups:
        print("⚠️ 分组来源为空或缺失，请先提供分组 json")
        sys.exit(1)
    if not col_mapping:
        print("⚠️ 列映射为空或缺失，请先提供列映射 json")
        sys.exit(2)

    if not os.path.exists(args.source_excel):
        print(f"❌ 找不到数据源 excel: {args.source_excel}")
        print("💡 请通过 --source-excel 指定实际数据源文件路径。")
        sys.exit(3)

    wb = openpyxl.load_workbook(args.source_excel, read_only=True, data_only=True)
    try:
        data = extract_data(wb, groups, col_mapping, sheet=args.sheet)
    finally:
        wb.close()

    result = {"data": data}

    os.makedirs(os.path.dirname(os.path.abspath(args.output)) or ".", exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"数据源: {args.source_excel}")
    print(f"groups 共 {len(groups)} 组，列映射源列 {len(col_mapping)} 个")
    for gname in groups:
        cols = data.get(gname, {})
        print(f"  {gname}: {len(cols)} 个目标列")
    print(f"结果已写入: {args.output}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"错误: {e}", file=sys.stderr)
        sys.exit(1)
