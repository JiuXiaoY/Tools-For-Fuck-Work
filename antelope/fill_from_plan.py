# -*- coding: utf-8 -*-
"""
fill_from_plan —— 按「填充计划(plan)」把数据写入模板副本，并自动清理多余空行。

按 11409 需求：数据项数目 m 以 plan.data 数组长度为准，**包含空数据（"" 也算一项）**，
例如 ["7aq...", "7aq...", "7aq...", "", "", "", "7aq..."] 视为 7 项。
模式判断（组共 n 行）：
    m == 0            → none（不填）
    m == n            → sequential（顺序填满 父+子）
    m == n - 1        → children_only（只填子体）
    m < n 且低于阈值  → cycle（循环铺满；cycle_threshold 为 null 则无条件允许）
    其他              → mismatch（不填，记入报告）

行号语义：plan.groups 存的是**实际行号（无偏移）**；填充到模板时应用偏移
    target 行 = 分组行 + (data_start_row − 1)
其中 data_start_row 来自 column_diff.json 的 settings.dataRow（唯一真源，plan 里带出）；
**读不到则报错退出，不做兜底**。
"""

import argparse
import json
import os
import sys
from pathlib import Path

import openpyxl

from common import setup_utf8, zcfg

# ═══════════════════════════════════════════════════════════════════════════════
# ⚙️ 默认运行配置（直接点 Run 时生效）—— 路径/配置项集中来自 zconfig.constant.py
# ═══════════════════════════════════════════════════════════════════════════════
DEFAULT_PLAN_FILE = zcfg.CFG_FILL_PLAN["default_plan_json"]   # 1. 你的 plan.json 路径
DEFAULT_OUTPUT_FILE = zcfg.OUTPUTS_DIR                        # 2. 完整输出文件路径(.xlsm)
DEFAULT_REPORT_FILE = zcfg.CFG_RUN["default_report_file"]     # 可选: "outputs/report.json"
DEFAULT_STRICT_SCOPE = zcfg.CFG_RUN["strict_scope"]           # 可选: True / False
# ═══════════════════════════════════════════════════════════════════════════════


def apply_column(
    ws,
    group_start_actual: int,
    group_end_actual: int,
    column: int,
    values,
    cycle_threshold,
    forced_mode=None,
):
    n = group_end_actual - group_start_actual + 1
    m = len(values)

    def _cycle_allowed():
        return cycle_threshold is None or m < cycle_threshold

    if m == 0:
        return "none", 0, "数据为空, 不填"

    # 手动指定模式时优先使用，跳过自动判断
    if forced_mode:
        if forced_mode == "cycle":
            target = list(range(group_start_actual, group_end_actual + 1))
            mode = "cycle"
        elif forced_mode == "children_only":
            target = list(range(group_start_actual + 1, group_end_actual + 1))
            mode = "children_only"
        elif forced_mode == "sequential":
            target = list(range(group_start_actual, group_end_actual + 1))
            mode = "sequential"
        else:
            return "mismatch", 0, f"无效的手动模式: {forced_mode!r}"

        if mode == "sequential" and m < n:
            # sequential 模式下值不足时按可填行数写
            target = target[: min(m, n)]
    else:
        if m == n:
            target = list(range(group_start_actual, group_end_actual + 1))
            mode = "sequential"
        elif m == n - 1:
            target = list(range(group_start_actual + 1, group_end_actual + 1))
            mode = "children_only"
        elif m < n and _cycle_allowed():
            target = list(range(group_start_actual, group_end_actual + 1))
            mode = "cycle"
        else:
            return "mismatch", 0, f"数据条数({m})与组行数({n})不匹配"

    filled = 0
    for i, row in enumerate(target):
        if mode == "cycle":
            value = values[i % m]
        else:
            if i >= len(values):
                break
            value = values[i]
        ws.cell(row=row, column=column).value = value
        filled += 1
    note = {
        "none": "不填",
        "sequential": f"顺序填写 {filled} 格(父+子)",
        "children_only": f"只填子体 {filled} 格(跳过父体)",
        "cycle": f"循环填写 {filled} 格(模式长度 {m})",
    }.get(mode, "")
    return mode, filled, note


def main():
    parser = argparse.ArgumentParser(
        description="按填充计划(plan.json)把数据写入模板副本(只读模板)。"
    )
    parser.add_argument(
        "plan",
        nargs="?",
        default=DEFAULT_PLAN_FILE,
        help="填充计划 JSON 文件路径",
    )
    parser.add_argument(
        "--output",
        "-o",
        default=DEFAULT_OUTPUT_FILE,
        help="输出 Excel 完整路径(含文件名.xlsm)",
    )
    parser.add_argument(
        "--report",
        default=DEFAULT_REPORT_FILE,
        help="填充报告 JSON 输出路径(可选)",
    )
    parser.add_argument(
        "--strict-scope",
        action="store_true",
        default=DEFAULT_STRICT_SCOPE,
        help="开启后: data 中出现的列若不在 col_scope 内则报错退出",
    )
    args = parser.parse_args()

    setup_utf8()

    if not args.plan or not os.path.exists(args.plan):
        print(f"❌ 找不到 plan 文件: {args.plan}")
        print("💡 请修改代码顶部的 DEFAULT_PLAN_FILE 为你的实际 JSON 文件路径。")
        sys.exit(1)

    try:
        plan = json.loads(Path(args.plan).read_text(encoding="utf-8"))
    except Exception as exc:
        print(f"无法读取 plan 文件 {args.plan}: {exc}")
        sys.exit(1)

    template_file = plan.get("template_file")
    if not template_file:
        print("plan 缺少 template_file")
        sys.exit(2)

    out_file = args.output or plan.get("output_file") or _default_output(template_file)
    if os.path.isdir(out_file) or out_file.endswith(("/", "\\")):
        out_file = os.path.join(out_file, "result_filled.xlsm")

    data_start_row = plan.get("data_start_row")
    if data_start_row is None:
        print("❌ plan 缺少 data_start_row（数据起始行）")
        print("💡 该值来自 column_diff.json 的 settings.dataRow（唯一真源），请先运行 analysisXlsm + column_diff + build_fill_framework 重新生成 plan。")
        sys.exit(5)
    try:
        data_start_row = int(data_start_row)
    except (TypeError, ValueError):
        print(f"❌ plan 的 data_start_row 非法: {data_start_row!r}")
        sys.exit(5)
    col_scope = set(plan.get("col_scope") or [])
    groups = plan.get("groups") or {}
    data = plan.get("data") or {}
    cycle_threshold = plan.get("cycle_threshold")
    mode_customise = plan.get("mode_customise") or {}

    if not os.path.exists(template_file):
        print(f"模板文件不存在: {template_file}")
        sys.exit(3)
    wb = openpyxl.load_workbook(template_file, keep_vba=True, data_only=False)
    # 主工作表名随国家配置变化（fr=Modèle / de=Vorlage），找不到则退回第一个工作表
    model_sheet = zcfg.SHEET_NAMES[zcfg.COUNTRY]["model"]
    ws_name = model_sheet if model_sheet in wb.sheetnames else wb.sheetnames[0]
    ws = wb[ws_name]

    offset = data_start_row - 1
    report_entries = []
    total_filled = 0
    out_of_scope = []
    max_filled_row = 0  # 记录所有分组覆盖到的最大实际行号

    # ---- 遍历每组每列 ----
    for gname, spec in groups.items():
        spec = str(spec).strip()
        if "&" not in spec:
            print(f"[skip] 组 {gname} 的规格无效: {spec!r}")
            continue
        try:
            start_w, end_w = map(int, spec.split("&"))
        except ValueError:
            print(f"[skip] 组 {gname} 的规格无效: {spec!r}")
            continue
        start_actual = start_w + offset
        end_actual = end_w + offset
        if start_actual < 1 or end_actual < start_actual:
            print(f"[skip] 组 {gname} 换算后行范围非法: {start_actual}..{end_actual}")
            continue

        # 更新最大已填行
        if end_actual > max_filled_row:
            max_filled_row = end_actual

        col_data = data.get(gname) or {}
        for col_str, values in col_data.items():
            try:
                col = int(col_str)
            except ValueError:
                print(f"[skip] 组 {gname} 列号非法: {col_str!r}")
                continue
            if col not in col_scope:
                out_of_scope.append((gname, col))
                if args.strict_scope:
                    print(f"[error] 列 {col} 不在 col_scope 内(组 {gname}), 开启 --strict-scope 报错退出。")
                    sys.exit(4)
                print(f"[warn] 列 {col} 不在 col_scope 内(组 {gname}), 跳过不填。")
                continue
            forced_mode = mode_customise.get(col_str)
            mode, filled, note = apply_column(
                ws,
                start_actual,
                end_actual,
                col,
                values,
                cycle_threshold,
                forced_mode=forced_mode,
            )
            total_filled += filled
            report_entries.append(
                {
                    "group": gname,
                    "rows": [start_actual, end_actual],
                    "column": col,
                    "mode": mode,
                    "filled": filled,
                    "count": len(values),
                    "note": note,
                }
            )
            print(
                f"  {gname:8} 列{col:<4} 模式={mode:<13} 填={filled:<3} 条数={len(values)}  {note}"
            )

    # ---- 整行删除：清理多余的模板历史行 ----
    if max_filled_row >= data_start_row:
        current_max_row = ws.max_row
        if current_max_row > max_filled_row:
            rows_to_delete = current_max_row - max_filled_row
            ws.delete_rows(max_filled_row + 1, rows_to_delete)
            print(f"🧹 已删除模板多余行: 第 {max_filled_row + 1} 行至第 {current_max_row} 行 (共整行删除 {rows_to_delete} 行)")

    # ---- 保存副本 ----
    out_dir = os.path.dirname(os.path.abspath(out_file))
    os.makedirs(out_dir, exist_ok=True)
    wb.save(out_file)

    # ---- 报告 ----
    report = {
        "template_file": template_file,
        "output_file": out_file,
        "data_start_row": data_start_row,
        "col_scope": sorted(col_scope),
        "groups": groups,
        "cycle_threshold": cycle_threshold,
        "total_filled": total_filled,
        "max_data_row": max_filled_row,
        "entries": report_entries,
        "out_of_scope_skipped": out_of_scope,
    }
    print(f"\n✅ 已写出副本: {out_file}")
    print(f"共填写 {total_filled} 个单元格，保留至第 {max_filled_row} 行。")

    if args.report:
        Path(args.report).write_text(
            json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(f"已写出报告: {args.report}")


def _default_output(template_file: str) -> str:
    base = os.path.splitext(os.path.basename(template_file))[0]
    return os.path.join(os.path.dirname(os.path.abspath(template_file)), f"{base}_filled.xlsm")


if __name__ == "__main__":
    main()