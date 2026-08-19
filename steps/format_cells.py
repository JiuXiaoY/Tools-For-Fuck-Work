"""Format cells: row height, alignment, column widths, formulas."""

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from core import PipelineContext, PipelineStep


class FormatCellsStep(PipelineStep):
    name = "format_cells"
    description = "Apply row height, alignment, column widths, formulas"
    requires = ("calc_price",)

    def run(self, ctx: PipelineContext) -> PipelineContext:
        cfg = self.config
        ws = ctx.worksheet

        # ── column widths ──
        letters_1_3 = (get_column_letter(1), get_column_letter(2), get_column_letter(3))
        for l in letters_1_3:
            ws.column_dimensions[l].width = cfg.col_width_1_3
        ws.column_dimensions[get_column_letter(4)].width = cfg.col_width_4
        ctx.log(f"Col 1-3 width={cfg.col_width_1_3}, col 4 width={cfg.col_width_4}")

        # ── row height + alignment + formula ──
        # Reuse a single Alignment object (openpyxl allows sharing style refs).
        align = Alignment(horizontal=cfg.cell_h_align, vertical=cfg.cell_v_align)
        max_col = ws.max_column
        col_d_letter = get_column_letter(4)
        n = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col):
            r = row[0].row
            ws.row_dimensions[r].height = cfg.row_height
            # Column 5 formula (col E = index 4 in the row tuple).
            if cfg.col_5_formula:
                row[4].value = f"=LEN({col_d_letter}{r})"
            # Apply alignment to every cell in the row (same style object reused).
            for cell in row:
                cell.alignment = align
            n += 1

        ctx.log(f"Formatted {n} rows: h={cfg.row_height}, align={cfg.cell_h_align}/{cfg.cell_v_align}")
        if cfg.col_5_formula:
            ctx.log(f"Column 5: =LEN({col_d_letter}) formula applied")
        return ctx
