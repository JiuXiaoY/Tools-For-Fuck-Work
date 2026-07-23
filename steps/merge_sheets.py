"""Merge all sheets in a workbook into one, migrating images and styles."""

import re
from copy import copy

from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor

from core import PipelineContext, PipelineStep

_DRAWING_RE = re.compile(r"^xl/drawings/drawing\d+\.xml$")


def _copy_style(src, dst) -> None:
    if not src.has_style:
        return
    for attr in ("font", "border", "fill", "number_format", "protection", "alignment"):
        setattr(dst, attr, copy(getattr(src, attr)))


class MergeSheetsStep(PipelineStep):
    name = "merge_sheets"
    description = "Merge all sheets into one, preserving images and styles"

    def run(self, ctx: PipelineContext) -> PipelineContext:
        names = ctx.workbook.sheetnames
        if len(names) <= 1:
            ctx.log("Single sheet, nothing to merge")
            return ctx

        target = ctx.worksheet
        nxt = target.max_row + 1
        copied = 0
        total_imgs = 0

        for name in names:
            if name == target.title:
                continue
            src = ctx.workbook[name]
            rows = src.max_row
            if rows == 0:
                ctx.log(f"Sheet [{name}]: empty, skip")
                continue

            offset = nxt - 1

            for i, sr in enumerate(range(1, rows + 1)):
                tr = nxt + i
                for c in range(1, src.max_column + 1):
                    sc = src.cell(row=sr, column=c)
                    tc = target.cell(row=tr, column=c)
                    tc.value = sc.value
                    _copy_style(sc, tc)
                if src.row_dimensions[sr].height:
                    target.row_dimensions[tr].height = src.row_dimensions[sr].height

            copied += rows

            # Copy images with row offset — same approach as Preprocess
            for img in src._images:
                a = img.anchor
                if not isinstance(a, (OneCellAnchor, TwoCellAnchor)):
                    continue

                new_img = copy(img)
                new_anchor = copy(a)
                new_anchor._from = copy(new_anchor._from)
                new_anchor._from.row += offset
                if isinstance(new_anchor, TwoCellAnchor) and new_anchor.to is not None:
                    new_anchor.to = copy(new_anchor.to)
                    new_anchor.to.row += offset

                new_img.anchor = new_anchor
                target.add_image(new_img)
                total_imgs += 1

            for mr in src.merged_cells.ranges:
                parts = str(mr).split(":")
                np = []
                for p in parts:
                    cl = p.rstrip("0123456789")
                    rn = int(p[len(cl):]) + offset
                    np.append(f"{cl}{rn}")
                try:
                    target.merge_cells(":".join(np))
                except Exception:
                    pass

            ctx.log(f"Sheet [{name}]: {rows} rows, {len(src._images)} images")
            nxt += rows

        for name in list(ctx.workbook.sheetnames):
            if name != target.title:
                del ctx.workbook[name]

        assets = ctx.metadata.get("assets", {})
        if assets:
            filtered = {k: v for k, v in assets.items() if not _DRAWING_RE.match(k)}
            ctx.metadata["assets"] = filtered
            skipped = len(assets) - len(filtered)
            if skipped:
                ctx.log(f"Skipped {skipped} source drawing file(s)")

        ctx.worksheet = target
        ctx.log(f"Merged {len(names)} sheets -> 1, {copied} data rows, {total_imgs} images")
        return ctx
