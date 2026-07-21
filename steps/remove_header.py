"""Remove header rows: if row-1 col-1 has no fill color, delete the row and shift images up."""

from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor

from core import PipelineContext, PipelineStep
from services import cell_has_fill


class RemoveHeaderStep(PipelineStep):
    name = "delete_header"
    description = "Remove header rows from all sheets based on fill color"

    def run(self, ctx: PipelineContext) -> PipelineContext:
        removed = 0
        for name in ctx.workbook.sheetnames:
            ws = ctx.workbook[name]
            if ws.max_row == 0:
                continue
            if cell_has_fill(ws.cell(row=1, column=1)):
                ctx.log(f"Sheet [{name}]: row 1 has fill, skip")
            else:
                ws.delete_rows(1)
                # openpyxl delete_rows does NOT shift image anchors — fix them
                for img in ws._images:
                    anchor = img.anchor
                    if isinstance(anchor, (OneCellAnchor, TwoCellAnchor)):
                        anchor._from.row = max(0, anchor._from.row - 1)
                        if isinstance(anchor, TwoCellAnchor) and anchor.to is not None:
                            anchor.to.row = max(0, anchor.to.row - 1)
                removed += 1
                ctx.log(f"Sheet [{name}]: header row removed, images adjusted")
        if removed == 0:
            ctx.log("All sheets: row 1 has fill, no header to remove")
        ctx.worksheet = ctx.workbook.active
        return ctx
