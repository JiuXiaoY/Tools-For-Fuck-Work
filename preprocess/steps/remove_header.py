"""Remove header rows: if row-1 col-1 has no fill, delete the row entirely (rebuild sheet)."""

from copy import copy

from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor
from openpyxl.workbook.workbook import Workbook

from services import cell_has_fill
from services.logger import get_logger

_log = get_logger("preprocess")


class RemoveHeaderPreStep:
    """Preprocessing step: remove header row from each sheet."""

    name = "remove_header"
    description = "Remove rows where col-1 has no fill, by rebuilding the sheet"

    def run(self, wb: Workbook, path: str) -> int:
        removed = 0
        for name in wb.sheetnames:
            ws = wb[name]
            if ws.max_row == 0:
                continue
            if cell_has_fill(ws.cell(row=1, column=1)):
                _log.info("  [%s] Sheet [%s]: row 1 has fill, skip", path, name)
                continue

            # ── Rebuild sheet skipping row 1 ──
            max_row = ws.max_row
            max_col = ws.max_column
            skip_row = 1

            ws_new = wb.create_sheet(title="_rm_header_tmp")

            # Copy rows (skip row 1), preserving row heights
            nr = 1
            for r in range(1, max_row + 1):
                if r == skip_row:
                    continue
                if ws.row_dimensions[r].height:
                    ws_new.row_dimensions[nr].height = ws.row_dimensions[r].height
                for c in range(1, max_col + 1):
                    src = ws.cell(row=r, column=c)
                    dst = ws_new.cell(row=nr, column=c)
                    dst.value = src.value
                    if src.has_style:
                        dst.font = copy(src.font)
                        dst.fill = copy(src.fill)
                        dst.border = copy(src.border)
                        dst.alignment = copy(src.alignment)
                        dst.number_format = src.number_format
                nr += 1

            # Copy images, shifting up by 1 row
            for img in ws._images:
                anchor = img.anchor
                old_row = None
                if isinstance(anchor, OneCellAnchor):
                    old_row = anchor._from.row + 1
                elif isinstance(anchor, TwoCellAnchor):
                    old_row = anchor._from.row + 1

                # 如果图片正好位于要删除的表头行上，则跳过
                if old_row == skip_row:
                    continue

                # 直接复制 openpyxl 的 Image 对象，避免 PIL 重新转码丢失清晰度和格式
                new_img = copy(img)
                new_anchor = copy(anchor)

                # 如果是依赖单元格定位的图片，进行平移操作
                if old_row is not None:
                    shift = -1
                    if isinstance(new_anchor, OneCellAnchor):
                        new_anchor._from = copy(new_anchor._from)
                        new_anchor._from.row += shift
                    elif isinstance(new_anchor, TwoCellAnchor):
                        new_anchor._from = copy(new_anchor._from)
                        new_anchor._from.row += shift
                        if new_anchor.to is not None:
                            new_anchor.to = copy(new_anchor.to)
                            new_anchor.to.row += shift

                new_img.anchor = new_anchor
                ws_new.add_image(new_img)

            # Copy merged cells (shift up by 1, skip those involving row 1)
            for merge_range in ws.merged_cells.ranges:
                if merge_range.min_row == skip_row or merge_range.max_row == skip_row:
                    continue
                ws_new.merge_cells(
                    min_row=merge_range.min_row - 1,
                    min_col=merge_range.min_col,
                    max_row=merge_range.max_row - 1,
                    max_col=merge_range.max_col,
                )

            # Copy column widths
            for col_letter, col_dim in ws.column_dimensions.items():
                if col_dim.width:
                    ws_new.column_dimensions[col_letter].width = col_dim.width

            # Replace old sheet
            ws_title = ws.title
            wb.remove(ws)
            ws_new.title = ws_title

            removed += 1
            _log.info("  [%s] Sheet [%s]: header row removed (sheet rebuilt)", path, name)

        if removed == 0:
            _log.info("  [%s] All sheets: row 1 has fill, nothing to remove", path)
        return removed
