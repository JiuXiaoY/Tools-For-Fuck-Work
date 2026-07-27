"""Remove rows where column J has no value — rebuild sheet."""

from copy import copy

from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor
from openpyxl.workbook.workbook import Workbook

from services.logger import get_logger

_log = get_logger("preprocess")


class RemoveEmptyJStep:
    """Preprocessing step: remove rows where col-J (column 10) is empty."""

    name = "remove_empty_j"
    description = "Remove rows where column J has no value, by rebuilding the sheet"

    def run(self, wb: Workbook, path: str) -> int:
        removed = 0
        for name in wb.sheetnames:
            ws = wb[name]
            if ws.max_row == 0:
                continue

            max_row = ws.max_row
            max_col = ws.max_column

            # ── Determine which rows to delete ──
            to_delete: set[int] = set()
            for r in range(1, max_row + 1):
                val = ws.cell(row=r, column=10).value
                if val is None or str(val).strip() == "":
                    to_delete.add(r)

            if not to_delete:
                _log.info("  [%s] Sheet [%s]: no empty J-column rows", path, name)
                continue

            # ── Rebuild sheet skipping deleted rows ──
            ws_new = wb.create_sheet(title="_rm_empty_j_tmp")

            # Copy row heights
            row_heights: dict[int, float] = {}
            for r in range(1, max_row + 1):
                if ws.row_dimensions[r].height:
                    row_heights[r] = ws.row_dimensions[r].height

            # Map old row → new row
            original_to_new: dict[int, int] = {}
            nr = 1
            for r in range(1, max_row + 1):
                if r in to_delete:
                    continue
                original_to_new[r] = nr
                nr += 1

            # Copy cells
            for r in range(1, max_row + 1):
                if r in to_delete:
                    continue
                nr = original_to_new[r]
                if r in row_heights:
                    ws_new.row_dimensions[nr].height = row_heights[r]
                for c in range(1, max_col + 1):
                    src = ws.cell(row=r, column=c)
                    dst = ws_new.cell(row=nr, column=c)
                    dst.value = src.value
                    if src.has_style:
                        dst.font = copy(src.font)
                        dst.fill = copy(src.fill)
                        dst.border = copy(src.border)
                        dst.alignment = copy(src.alignment)
                        dst.number_format = src.number_format

            # Copy images
            for img in ws._images:
                anchor = img.anchor
                old_row = None
                if isinstance(anchor, OneCellAnchor):
                    old_row = anchor._from.row + 1
                elif isinstance(anchor, TwoCellAnchor):
                    old_row = anchor._from.row + 1

                if old_row is not None and old_row in to_delete:
                    continue

                new_img = copy(img)
                new_anchor = copy(anchor)

                if old_row is not None and old_row in original_to_new:
                    shift = original_to_new[old_row] - old_row
                    if isinstance(new_anchor, OneCellAnchor):
                        new_anchor._from = copy(new_anchor._from)
                        new_anchor._from.row += shift
                    elif isinstance(new_anchor, TwoCellAnchor):
                        new_anchor._from = copy(new_anchor._from)
                        new_anchor._from.row += shift
                        if new_anchor.to is not None:
                            new_anchor.to = copy(new_anchor.to)
                            new_anchor.to.row += shift

                new_img.anchor = new_anchor
                ws_new.add_image(new_img)

            # Copy merged cells
            for merge_range in ws.merged_cells.ranges:
                top, left, bottom, right = merge_range.min_row, merge_range.min_col, merge_range.max_row, merge_range.max_col
                if top in to_delete or bottom in to_delete:
                    continue
                if top in original_to_new and bottom in original_to_new:
                    ws_new.merge_cells(
                        min_row=original_to_new[top],
                        min_col=left,
                        max_row=original_to_new[bottom],
                        max_col=right,
                    )

            # Copy column widths
            for col_letter, col_dim in ws.column_dimensions.items():
                if col_dim.width:
                    ws_new.column_dimensions[col_letter].width = col_dim.width

            # Replace old sheet
            ws_title = ws.title
            wb.remove(ws)
            ws_new.title = ws_title

            removed += len(to_delete)
            _log.info("  [%s] Sheet [%s]: %d rows with empty J-column removed (sheet rebuilt)",
                      path, name, len(to_delete))

        if removed == 0:
            _log.info("  [%s] All sheets: no empty J-column rows", path)
        return removed
