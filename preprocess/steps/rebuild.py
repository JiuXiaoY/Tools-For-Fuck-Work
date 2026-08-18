"""Shared worksheet rebuild utility used by preprocessing steps."""

from __future__ import annotations

from collections.abc import Callable
from copy import copy

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from services.images import image_anchor_row, shift_image_anchor

TEMP_SHEET_TITLE = "_preprocess_tmp"


def rebuild_sheet(
    workbook: Workbook,
    worksheet: Worksheet,
    keep_row: Callable[[int], bool],
) -> Worksheet:
    """Replace ``worksheet`` with a copy containing rows accepted by ``keep_row``."""
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    row_map: dict[int, int] = {}
    next_row = 1
    for row in range(1, max_row + 1):
        if keep_row(row):
            row_map[row] = next_row
            next_row += 1

    rebuilt = workbook.create_sheet(title=TEMP_SHEET_TITLE)
    for old_row, new_row in row_map.items():
        height = worksheet.row_dimensions[old_row].height
        if height is not None:
            rebuilt.row_dimensions[new_row].height = height
        for column in range(1, max_col + 1):
            source = worksheet.cell(row=old_row, column=column)
            target = rebuilt.cell(row=new_row, column=column, value=source.value)
            if source.has_style:
                target.font = copy(source.font)
                target.fill = copy(source.fill)
                target.border = copy(source.border)
                target.alignment = copy(source.alignment)
                target.number_format = source.number_format
                target.protection = copy(source.protection)

    for image in worksheet._images:
        anchor = image.anchor
        old_row = image_anchor_row(anchor)
        if old_row is None:
            continue
        if old_row not in row_map:
            continue
        row_shift = row_map[old_row] - old_row
        new_image = copy(image)
        new_image.anchor = shift_image_anchor(anchor, row_shift)
        rebuilt.add_image(new_image)

    for merged in worksheet.merged_cells.ranges:
        if merged.min_row in row_map and merged.max_row in row_map:
            rebuilt.merge_cells(
                min_row=row_map[merged.min_row],
                min_col=merged.min_col,
                max_row=row_map[merged.max_row],
                max_col=merged.max_col,
            )

    for column_letter, dimension in worksheet.column_dimensions.items():
        if dimension.width is not None:
            rebuilt.column_dimensions[column_letter].width = dimension.width

    original_title = worksheet.title
    original_index = workbook.index(worksheet)
    workbook.remove(worksheet)
    rebuilt.title = original_title
    workbook._sheets.remove(rebuilt)
    workbook._sheets.insert(original_index, rebuilt)
    return rebuilt
