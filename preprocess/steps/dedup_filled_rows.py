"""Deduplicate filled rows by column C value with distance threshold.

Only considers rows where col-A has a fill ("filled rows").
For each adjacent pair in the filled-row sequence:
    - If the rows between two filled rows <= close_gap (default 5, config
      preprocess_dedup_close_gap): delete the anchor row and all rows in
      between; the later row is kept and becomes the new anchor.
    - Otherwise, if col-C values are equal AND the row gap <= max_gap:
      delete the later row, keeping the anchor row.
    - Otherwise: the later row becomes the new anchor.
"""

from copy import copy

from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor
from openpyxl.workbook.workbook import Workbook

from services import cell_has_fill
from services.logger import get_logger

_log = get_logger("preprocess")


class DedupFilledRowsStep:
    """Preprocessing: deduplicate filled rows by col-C value."""

    name = "dedup_filled_rows"
    description = "Remove close filled rows (<=5 between) + dedup by col-C with max_gap, anchor-based"

    def run(self, wb: Workbook, path: str) -> int:
        from config import Config
        cfg = Config()
        max_gap = cfg.preprocess_dedup_max_gap
        close_gap = cfg.preprocess_dedup_close_gap

        ws = wb.active
        max_row = ws.max_row
        max_col = ws.max_column

        # ── Collect filled rows ──
        filled: list[tuple[int, object]] = []  # (row, col_C_value)
        for r in range(1, max_row + 1):
            if cell_has_fill(ws.cell(row=r, column=1)):
                c_val = ws.cell(row=r, column=3).value
                filled.append((r, c_val))

        if len(filled) < 2:
            _log.info("  [%s] Less than 2 filled rows, nothing to dedup", path)
            return 0

        # ── Determine which rows to delete (anchor-based) ──
        to_delete: set[int] = set()
        r_anchor, c_anchor = filled[0]

        for i in range(1, len(filled)):
            r_curr, c_curr = filled[i]
            between = (r_curr - r_anchor) - 1  # 两有色行之间夹的行数

            if between <= close_gap:
                # 间距很近(之间≤5行)：删除锚点行和之间的所有行，保留当前行作为新锚点
                to_delete.add(r_anchor)
                for mid_row in range(r_anchor + 1, r_curr):
                    to_delete.add(mid_row)
                r_anchor, c_anchor = r_curr, c_curr
            elif _same_value(c_anchor, c_curr) and (r_curr - r_anchor) <= max_gap:
                # C相同且间距在阈值内 → 删除后面的行，保持锚点不变
                to_delete.add(r_curr)
            else:
                # C不同或间距超阈值 → 当前行成为新锚点
                r_anchor, c_anchor = r_curr, c_curr

        if not to_delete:
            _log.info("  [%s] No duplicate filled rows to remove", path)
            return 0

        # ── Rebuild sheet: keep only non-deleted rows ──
        # Collect all row data (styles, values, images per row)
        row_data: list[list] = []  # list of cells per row to keep
        row_heights: dict[int, float] = {}  # original row heights
        original_to_new: dict[int, int] = {}  # old row → new row mapping

        # Read row heights
        for r in range(1, max_row + 1):
            if ws.row_dimensions[r].height:
                row_heights[r] = ws.row_dimensions[r].height

        new_row = 1
        for r in range(1, max_row + 1):
            if r in to_delete:
                continue
            original_to_new[r] = new_row
            new_row += 1

        # Create a new sheet to copy data into
        ws_new = wb.create_sheet(title="_dedup_tmp")
        for r in range(1, max_row + 1):
            if r in to_delete:
                continue
            nr = original_to_new[r]
            if nr in row_heights:
                ws_new.row_dimensions[nr].height = row_heights.get(r)
            for c in range(1, max_col + 1):
                src = ws.cell(row=r, column=c)
                dst = ws_new.cell(row=nr, column=c)
                dst.value = src.value
                if src.has_style:
                    dst.font = copy(src.font)
                    dst.fill = copy(src.fill)
                    dst.border = copy(src.border)
                    dst.alignment = copy(src.alignment)
                    dst.number_format = src.number_format

        # Copy images
        for img in ws._images:
            anchor = img.anchor
            old_row = None
            if isinstance(anchor, OneCellAnchor):
                old_row = anchor._from.row + 1
            elif isinstance(anchor, TwoCellAnchor):
                old_row = anchor._from.row + 1

            # 如果图片关联的行已经被去重删除，则跳过不复制
            if old_row is not None and old_row in to_delete:
                continue

            new_img = copy(img)
            new_anchor = copy(anchor)

            if old_row is not None and old_row in original_to_new:
                shift = original_to_new[old_row] - old_row
                if isinstance(new_anchor, OneCellAnchor):
                    new_anchor._from = copy(new_anchor._from)
                    new_anchor._from.row += shift
                elif isinstance(new_anchor, TwoCellAnchor):
                    new_anchor._from = copy(new_anchor._from)
                    new_anchor._from.row += shift
                    if new_anchor.to is not None:
                        new_anchor.to = copy(new_anchor.to)
                        new_anchor.to.row += shift

            new_img.anchor = new_anchor
            ws_new.add_image(new_img)

        # Copy merged cells
        for merge_range in ws.merged_cells.ranges:
            top, left, bottom, right = merge_range.min_row, merge_range.min_col, merge_range.max_row, merge_range.max_col
            if top in to_delete or bottom in to_delete:
                continue
            if top in original_to_new and bottom in original_to_new:
                ws_new.merge_cells(
                    min_row=original_to_new[top],
                    min_col=left,
                    max_row=original_to_new[bottom],
                    max_col=right,
                )

        # Copy column widths
        for col_letter, col_dim in ws.column_dimensions.items():
            if col_dim.width:
                ws_new.column_dimensions[col_letter].width = col_dim.width

        # Replace old sheet with new one
        ws_title = ws.title
        wb.remove(ws)
        ws_new.title = ws_title

        _log.info("  [%s] Dedup: %d filled rows, %d deleted (sheet rebuilt)", path, len(filled), len(to_delete))
        return len(to_delete)


def _same_value(a: object, b: object) -> bool:
    """Check if two cell values are equal (handle None)."""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return str(a).strip() == str(b).strip()
