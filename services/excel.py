"""Excel workbook I/O: load, save, merge (first sheet only from each file)."""

from __future__ import annotations

import logging
from copy import copy as copy_style
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from services.images import clone_image

logger = logging.getLogger(__name__)


def load(path: Path) -> Workbook:
    return load_workbook(path)


def save(wb: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _copy_style(src, dst) -> None:
    if not src.has_style:
        return
    for attr in ("font", "border", "fill", "number_format", "protection", "alignment"):
        setattr(dst, attr, copy_style(getattr(src, attr)))


def merge_workbooks(paths: list[Path]) -> dict:
    """Merge first sheet of each workbook into one, returning {workbook, files, images}."""
    if not paths:
        raise ValueError("empty path list")

    target = load(paths[0])
    for name in list(target.sheetnames):
        if name != target.active.title:
            del target[name]

    total = len(target.active._images)
    print(f"  [merge] {paths[0].name}: {total}/{total} images kept")

    for fp in paths[1:]:
        src = load(fp)
        src_ws = src.active
        new_name = src_ws.title
        if new_name in target.sheetnames:
            base, n = new_name, 2
            while f"{base}_{n}" in target.sheetnames:
                n += 1
            new_name = f"{base}_{n}"

        tws = target.create_sheet(title=new_name)

        for col_letter, cd in src_ws.column_dimensions.items():
            if cd.width:
                tws.column_dimensions[col_letter].width = cd.width

        for row in src_ws.iter_rows():
            rn = row[0].row
            rd = src_ws.row_dimensions[rn]
            if rd.height:
                tws.row_dimensions[rn].height = rd.height
            for cell in row:
                tc = tws.cell(row=cell.row, column=cell.column)
                tc.value = cell.value
                _copy_style(cell, tc)

        for mr in src_ws.merged_cells.ranges:
            try:
                tws.merge_cells(str(mr))
            except Exception:
                pass

        img_count = 0
        for img in src_ws._images:
            try:
                if clone_image(img, tws):
                    img_count += 1
            except Exception as exc:
                logger.warning("Image clone failed in %s: %s", fp.name, exc)
        total += img_count
        print(f"  [merge] {fp.name}: {img_count}/{len(src_ws._images)} images copied")

    return {"workbook": target, "images": total}
