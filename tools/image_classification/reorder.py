"""Reorder image links in Excel: copy size charts to the end and mark non-conforming positions red.

Data source: O-W columns (15-23) of outputs/{date}v1.xlsx

Multi-threaded: rows split across threads, each thread processes independently.
Results merged, then applied to Excel in single thread.

Usage:
    python tools/image_classification/reorder.py
"""

from __future__ import annotations

import sys
import time
import uuid
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import requests
from requests.adapters import HTTPAdapter
from openpyxl.styles import PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from config import Config
from services.logger import get_logger

BASE = Path(__file__).resolve().parent
TEMP_DIR = BASE / "images_awaiting"
_log = get_logger("image_reorder")

COL_START = 15
COL_END = 23
# 将并发数提升至 16，充分利用网络和 CPU，如果被服务器拦截可适当调小
WORKERS = 5

# 使用 threading.local() 保证每个线程有独立的 Session 和连接池
_thread_local = threading.local()


def _get_session() -> requests.Session:
    """获取线程安全的 Session，并放大连接池"""
    if not hasattr(_thread_local, "session"):
        sess = requests.Session()
        # 放大底层连接池，与并发线程数匹配，避免排队干等
        adapter = HTTPAdapter(pool_connections=WORKERS, pool_maxsize=WORKERS)
        sess.mount('http://', adapter)
        sess.mount('https://', adapter)
        _thread_local.session = sess
    return _thread_local.session


def _download(url: str) -> Path | None:
    """下载图片：使用 UUID 防止多线程环境下的文件系统竞争"""
    TEMP_DIR.mkdir(parents=True, exist_ok=True)

    # 提取扩展名，默认为 jpg，使用 UUID 保证绝对唯一性避免多线程覆写冲突
    ext = url.split("/")[-1].split("?")[0].split(".")[-1]
    if not ext or len(ext) > 4:
        ext = "jpg"
    out = TEMP_DIR / f"{uuid.uuid4().hex}.{ext}"

    sess = _get_session()
    for attempt in range(1, 4):
        try:
            # 缩短超时时间，遇到卡死的图片尽快重试
            resp = sess.get(url, timeout=100)
            resp.raise_for_status()
            out.write_bytes(resp.content)
            return out
        except Exception:
            if attempt < 3:
                time.sleep(1)
    return None


def _is_size_chart(img_path: Path, cfg: Config) -> bool:
    from tools.image_classification.classify import _detect_ocr, _detect_opencv, _detect_heuristic
    mode = cfg.img_classify_mode
    scores: dict[str, float] = {}
    if mode in ("heuristic", "all"):
        scores["heuristic"] = _detect_heuristic(img_path)
    if mode in ("ocr", "all"):
        scores["ocr"] = _detect_ocr(img_path, cfg)
    if mode in ("opencv", "all"):
        scores["opencv"] = _detect_opencv(img_path, cfg)
    if mode == "all":
        valid = {k: v for k, v in scores.items() if v > 0.0}
        if not valid:
            return False
        yes = sum(1 for v in valid.values() if v >= 0.5)
        avg = sum(valid.values()) / len(valid)
        return yes >= 2 or avg >= 0.6
    return scores.get(mode, 0.0) >= 0.5


def main() -> None:
    cfg = Config()
    date_str = _resolve_date(cfg)
    src = Path(__file__).resolve().parent.parent.parent / cfg.out_dir / f"{date_str}v1.xlsx"

    if not src.exists():
        _log.error("Source not found: %s", src)
        return

    import openpyxl
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb.active
    max_row = ws.max_row

    _log.info("Source: %s  Rows: %d  Range: O(%d)-W(%d)  Mode: %s  Workers: %d",
              src.name, max_row, COL_START, COL_END, cfg.img_classify_mode, WORKERS)

    # ── Collect all row data (fast, read-only) ──
    all_rows: dict[int, list[str | None]] = {}
    for r in range(1, max_row + 1):
        links: list[str | None] = []
        for c in range(COL_START, COL_END + 1):
            val = ws.cell(row=r, column=c).value
            if val is None or str(val).strip() == "":
                links.append(None)
            else:
                links.append(str(val).strip())
        all_rows[r] = links
    wb.close()

    # ── Split rows across threads ──
    row_nums = list(all_rows.keys())
    chunk_size = max(1, len(row_nums) // WORKERS)
    chunks = [row_nums[i:i + chunk_size] for i in range(0, len(row_nums), chunk_size)]

    _log.info("Processing %d rows in %d chunks", len(row_nums), len(chunks))

    results: list[dict] = []
    # 使用线程池并发执行，开启 WORKERS 数量的线程
    with ThreadPoolExecutor(max_workers=WORKERS, thread_name_prefix="Worker") as pool:
        fut_to_chunk = {}
        for ci, chunk in enumerate(chunks):
            fut = pool.submit(_process_chunk, chunk, all_rows, cfg, ci + 1)
            fut_to_chunk[fut] = ci

        for fut in as_completed(fut_to_chunk):
            chunk_results = fut.result()
            results.extend(chunk_results)

    # Sort by row
    results.sort(key=lambda x: x["row"])

    # ── Apply to Excel ──
    wb = openpyxl.load_workbook(src)
    ws = wb.active

    def _mark_red(ws, row: int, col: int) -> None:
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid", patternType="solid")
        from copy import copy
        if not cell.has_style:
            cell.font = copy(cell.font)

    rows_red = 0
    for res in results:
        r = res["row"]
        num_links = res["num_links"]

        if res["all_red"]:
            # 全部找不到尺码表，整行图片标红
            for c in range(COL_START, COL_START + num_links):
                _mark_red(ws, r, c)
            rows_red += 1
            continue

        size_chart_idx = res.get("size_chart_idx")
        if size_chart_idx is not None:
            actual_col = COL_START + size_chart_idx
            last_col = COL_START + num_links - 1

            # 如果尺码表已经是在当前行的最后一位，说明排版没问题，跳过
            if actual_col == last_col:
                continue

            # 不移动原图，直接在末尾追加（复制）一张尺码表
            dst_col = last_col + 1
            # 越界保护：如果追加的列超过了 W 列(COL_END)，则只能强制覆盖在 W 列上
            if dst_col > COL_END:
                dst_col = COL_END

            src_cell = ws.cell(row=r, column=actual_col)
            dst_cell = ws.cell(row=r, column=dst_col)

            # 复制尺码表的值和链接到末尾
            dst_cell.value = src_cell.value
            dst_cell.hyperlink = src_cell.hyperlink

            # 将原来位置的尺码表标红
            _mark_red(ws, r, actual_col)
            rows_red += 1

    wb.save(src)
    wb.close()
    _log.info("Done: %d rows processed, %d rows with red marks (duplicated/missing)", len(results), rows_red)


def _process_chunk(row_nums: list[int], all_rows: dict, cfg: Config, chunk_id: int) -> list[dict]:
    """Process a chunk of rows in one thread, lazy scan per row."""
    # 获取当前执行该函数的线程名称
    thread_name = threading.current_thread().name
    _log.info("[%s] Started processing Chunk %d (%d rows)", thread_name, chunk_id, len(row_nums))

    results = []
    start_col = 0 # 改为 0-based index，代表在 links 数组中的位置
    for i, r in enumerate(row_nums):
        links: list[str] = []
        for val in all_rows[r]:
            if val is None:
                break
            links.append(val)

        if not links:
            results.append({"row": r, "num_links": 0, "all_red": False})
            continue

        num_links = len(links)
        si = min(start_col, num_links - 1)
        found_idx: int | None = None

        # 懒加载验证
        def _check(idx: int) -> bool | None:
            url = links[idx]
            p = _download(url)
            if p is None:
                return None
            try:
                result = _is_size_chart(p, cfg)
                return result
            except Exception:
                return False
            finally:
                try: p.unlink()
                except OSError: pass

        r0 = _check(si)
        if r0 is None: # 下载失败或网络错误
            results.append({"row": r, "num_links": num_links, "all_red": True})
            start_col = 0
            continue

        if r0:
            found_idx = si
        else:
            # 向后查找
            for j in range(si + 1, num_links):
                rj = _check(j)
                if rj is None:
                    break
                if rj:
                    found_idx = j
                    break
            # 如果向后没找到，向前查找
            if found_idx is None:
                for j in range(0, si):
                    rj = _check(j)
                    if rj is None:
                        break
                    if rj:
                        found_idx = j
                        break

        if found_idx is None:
            # 这一行全部找不到尺码表
            results.append({"row": r, "num_links": num_links, "all_red": True})
            start_col = 0
        else:
            # 找到了尺码表，返回其索引，交给 Excel 处理线程去追加复制
            results.append({"row": r, "num_links": num_links, "size_chart_idx": found_idx, "all_red": False})
            start_col = found_idx

        if (i + 1) % 50 == 0:
            _log.info("  [%s] Chunk %d: %d/%d rows processed", thread_name, chunk_id, i + 1, len(row_nums))

    _log.info("[%s] Completed Chunk %d", thread_name, chunk_id)
    return results


def _resolve_date(cfg: Config) -> str:
    raw = cfg.date_override.strip()
    if raw and len(raw) == 6:
        try:
            month = int(raw[2:4])
            day = int(raw[4:6])
            return f"{month}.{day}"
        except ValueError:
            pass
    from datetime import datetime
    today = datetime.now()
    return f"{today.month}.{today.day}"


if __name__ == "__main__":
    main()