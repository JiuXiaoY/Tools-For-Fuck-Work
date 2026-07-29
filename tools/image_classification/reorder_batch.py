"""Batch reorder image links by colored first-cell groups.

Difference from reorder.py:
  - Group rows by colored first-cell (column A):
    rows [colored_1, colored_2) = one batch (same product family)
  - Only scan the FIRST data row per batch to locate the size chart.
  - All subsequent rows in the same batch REUSE the same column index.
  - This avoids repeatedly downloading + classifying identical layouts.

Data source: O-W columns (15-23) of outputs/{date}v1.xlsx
Detection:   Reuses classify.py three engines (heuristic / ocr / opencv)

Usage:
    python tools/image_classification/reorder_batch.py
"""

from __future__ import annotations

import sys
import time
import uuid
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import requests
from requests.adapters import HTTPAdapter
from openpyxl.styles import PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from config import Config
from services.logger import get_logger

BASE = Path(__file__).resolve().parent
TEMP_DIR = BASE / "images_awaiting"
_log = get_logger("image_reorder_batch")

COL_START = 15          # O
COL_END   = 23          # W
COL_FIRST = 1           # A — used to detect coloured batch markers
WORKERS   = 6

_thread_local = threading.local()


def _get_session() -> requests.Session:
    if not hasattr(_thread_local, "session"):
        sess = requests.Session()
        adapter = HTTPAdapter(pool_connections=WORKERS, pool_maxsize=WORKERS)
        sess.mount("http://", adapter)
        sess.mount("https://", adapter)
        _thread_local.session = sess
    return _thread_local.session


def _download(url: str) -> Path | None:
    TEMP_DIR.mkdir(parents=True, exist_ok=True)
    ext = url.split("/")[-1].split("?")[0].split(".")[-1]
    if not ext or len(ext) > 4:
        ext = "jpg"
    out = TEMP_DIR / f"{uuid.uuid4().hex}.{ext}"
    sess = _get_session()
    for attempt in range(1, 4):
        try:
            resp = sess.get(url, timeout=100)
            resp.raise_for_status()
            out.write_bytes(resp.content)
            return out
        except Exception:
            if attempt < 3:
                time.sleep(1)
    return None


def _is_size_chart(img_path: Path, cfg: Config) -> bool:
    from tools.image_classification.classify import _detect_ocr, _detect_opencv, _detect_heuristic
    mode = cfg.img_classify_mode
    scores: dict[str, float] = {}
    if mode in ("heuristic", "all"):
        scores["heuristic"] = _detect_heuristic(img_path)
    if mode in ("ocr", "all"):
        scores["ocr"] = _detect_ocr(img_path, cfg)
    if mode in ("opencv", "all"):
        scores["opencv"] = _detect_opencv(img_path, cfg)
    if mode == "all":
        valid = {k: v for k, v in scores.items() if v > 0.0}
        if not valid:
            return False
        yes = sum(1 for v in valid.values() if v >= 0.5)
        avg = sum(valid.values()) / len(valid)
        return yes >= 2 or avg >= 0.6
    return scores.get(mode, 0.0) >= 0.5


def _has_fill(ws, row: int, col: int = COL_FIRST) -> bool:
    """Check if a cell has a fill (background colour)."""
    cell = ws.cell(row=row, column=col)
    fill = cell.fill
    if fill and fill.start_color and fill.start_color.rgb:
        rgb = str(fill.start_color.rgb)
        # "00000000" = no fill; fgColor with no colour = patternType=None
        if rgb.upper() not in ("00000000", "FFFFFFFF", "") and fill.patternType:
            return True
    return False


def _build_batches(ws, max_row: int) -> list[tuple[int, int]]:
    """Scan column A and return [(start_row, end_row), …] batches.

    Each batch is the range of rows *between* two consecutive coloured
    cells in column A.  The coloured rows themselves are SKIPPED.
    """
    markers = []
    for r in range(1, max_row + 1):
        if _has_fill(ws, r):
            markers.append(r)

    batches: list[tuple[int, int]] = []
    for i in range(len(markers) - 1):
        start = markers[i] + 1
        end   = markers[i + 1] - 1
        if start <= end:
            batches.append((start, end))

    # Last marker → end of sheet
    if markers:
        start = markers[-1] + 1
        if start <= max_row:
            batches.append((start, max_row))

    return batches


def _resolve_date(cfg: Config) -> str:
    raw = cfg.date_override.strip()
    if raw and len(raw) == 6:
        try:
            month = int(raw[2:4])
            day   = int(raw[4:6])
            return f"{month}.{day}"
        except ValueError:
            pass
    from datetime import datetime
    today = datetime.now()
    return f"{today.month}.{today.day}"


def main() -> None:
    cfg = Config()
    date_str = _resolve_date(cfg)
    src = Path(__file__).resolve().parent.parent.parent / cfg.out_dir / f"{date_str}v1.xlsx"

    if not src.exists():
        _log.error("Source not found: %s", src)
        return

    import openpyxl
    wb = openpyxl.load_workbook(src)
    ws = wb.active
    max_row = ws.max_row

    # ── 1. Build batches by coloured A-column cells ──
    batches = _build_batches(ws, max_row)
    _log.info("Source: %s  MaxRow: %d  Batches: %d  Range: O(%d)-W(%d)  Mode: %s",
              src.name, max_row, len(batches), COL_START, COL_END, cfg.img_classify_mode)

    if not batches:
        _log.warning("No coloured markers found in column A — nothing to do.")
        wb.close()
        return

    # ── 2. Collect image links for all rows ──
    all_rows: dict[int, list[str | None]] = {}
    for r in range(1, max_row + 1):
        links: list[str | None] = []
        for c in range(COL_START, COL_END + 1):
            val = ws.cell(row=r, column=c).value
            if val is None or str(val).strip() == "":
                links.append(None)
            else:
                links.append(str(val).strip())
        all_rows[r] = links
    wb.close()

    # ── 3. Process batches in parallel ──
    # Each batch is an atomic unit: same leader row, same size-chart position.
    # Assign full batches to workers (not individual rows) to avoid
    # redundant leader scans when a batch spans chunk boundaries.
    _log.info("Processing %d batches across %d workers", len(batches), WORKERS)

    batch_list = list(batches)  # (start, end) tuples
    chunk_size = max(1, len(batch_list) // WORKERS)
    chunks = [batch_list[i:i + chunk_size] for i in range(0, len(batch_list), chunk_size)]

    results: list[dict] = []
    with ThreadPoolExecutor(max_workers=WORKERS, thread_name_prefix="Batch") as pool:
        futs = []
        for ci, chunk_batches in enumerate(chunks):
            fut = pool.submit(_process_batches, chunk_batches, all_rows, cfg, ci + 1)
            futs.append(fut)
        for fut in as_completed(futs):
            chunk_results = fut.result()
            results.extend(chunk_results)

    results.sort(key=lambda x: x["row"])

    # ── 4. Apply to Excel ──
    wb = openpyxl.load_workbook(src)
    ws = wb.active

    def _mark_red(ws, row: int, col: int) -> None:
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC",
                                fill_type="solid", patternType="solid")
        from copy import copy
        if not cell.has_style:
            cell.font = copy(cell.font)

    def _copy_cell(ws, row: int, src_col: int, dst_col: int) -> None:
        src = ws.cell(row=row, column=src_col)
        dst = ws.cell(row=row, column=dst_col)
        dst.value = src.value
        dst.hyperlink = src.hyperlink

    rows_red = 0
    for res in results:
        r = res["row"]
        num_links = res["num_links"]

        if res["all_red"]:
            for c in range(COL_START, COL_START + num_links):
                _mark_red(ws, r, c)
            rows_red += 1
            continue

        size_chart_idx = res.get("size_chart_idx")
        if size_chart_idx is not None:
            actual_col = COL_START + size_chart_idx
            last_col   = COL_START + num_links - 1

            # If size_chart_idx exceeds current row's link count, mark red
            if size_chart_idx >= num_links:
                for c in range(COL_START, COL_START + num_links):
                    _mark_red(ws, r, c)
                rows_red += 1
                continue

            if cfg.img_reorder_mode == "move_dual":
                # Read all cell values + hyperlinks in this row
                cells = []
                for c in range(COL_START, COL_START + num_links):
                    cell = ws.cell(row=r, column=c)
                    cells.append({
                        "value": cell.value,
                        "hyperlink": cell.hyperlink,
                    })

                chart = cells[size_chart_idx]
                # Build new order: non-chart items first, then 2 copies of chart
                rest = [c for i, c in enumerate(cells) if i != size_chart_idx]
                ordered = rest + [chart, chart]

                # Write back, no gaps (within O-W range)
                for i, cdata in enumerate(ordered):
                    col = COL_START + i
                    if col > COL_END:
                        break
                    dst = ws.cell(row=r, column=col)
                    dst.value = cdata["value"]
                    dst.hyperlink = cdata.get("hyperlink")

                # Clear any remaining columns in O-W range
                for c in range(COL_START + len(ordered), COL_END + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.value = None
                    cell.hyperlink = None

            elif cfg.img_reorder_mode == "inline_dual":
                # Read all cell values in this row
                cells = []
                for c in range(COL_START, COL_START + num_links):
                    cell = ws.cell(row=r, column=c)
                    cells.append({
                        "value": cell.value,
                        "hyperlink": cell.hyperlink,
                    })

                chart = cells[size_chart_idx]
                # Insert a copy right after original, then append another at tail
                ordered = list(cells)
                ordered.insert(size_chart_idx + 1, chart)
                ordered.append(chart)

                # Truncate from end if exceeds column range
                max_cols = COL_END - COL_START + 1
                if len(ordered) > max_cols:
                    ordered = ordered[:max_cols]

                # Write back, no gaps
                for i, cdata in enumerate(ordered):
                    col = COL_START + i
                    dst = ws.cell(row=r, column=col)
                    dst.value = cdata["value"]
                    dst.hyperlink = cdata.get("hyperlink")

                # Clear remaining columns
                for c in range(COL_START + len(ordered), COL_END + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.value = None
                    cell.hyperlink = None

            else:  # "copy_single" (default)
                if actual_col == last_col:
                    continue

                dst_col = last_col + 1
                if dst_col > COL_END:
                    dst_col = COL_END

                _copy_cell(ws, r, actual_col, dst_col)
                _mark_red(ws, r, actual_col)

            rows_red += 1

    wb.save(src)
    wb.close()
    _log.info("Done: %d rows processed, %d rows with red marks", len(results), rows_red)


def _process_batches(batches: list[tuple[int, int]],
                     all_rows: dict,
                     cfg: Config,
                     chunk_id: int) -> list[dict]:
    """Process a list of (start, end) batch tuples."""
    thread_name = threading.current_thread().name
    _log.info("[%s] Chunk %d started (%d batches)", thread_name, chunk_id, len(batches))

    results = []

    for bi, (batch_start, batch_end) in enumerate(batches):
        leader = batch_start
        # Scan leader once per batch
        idx = _scan_leader(leader, all_rows, cfg)
        _log.info("  [%s] Batch leader row %d → size_chart_idx=%s",
                  thread_name, leader, idx)

        for r in range(batch_start, batch_end + 1):
            links: list[str] = []
            for val in all_rows[r]:
                if val is None:
                    break
                links.append(val)
            num_links = len(links)

            if num_links == 0:
                results.append({"row": r, "num_links": 0, "all_red": False})
                continue

            if idx is None:
                results.append({"row": r, "num_links": num_links, "all_red": True})
            else:
                results.append({
                    "row": r,
                    "num_links": num_links,
                    "size_chart_idx": idx,
                    "all_red": False,
                })

        if (bi + 1) % 20 == 0:
            _log.info("  [%s] Chunk %d: %d/%d batches", thread_name, chunk_id, bi + 1, len(batches))

    _log.info("[%s] Chunk %d completed (%d batches)", thread_name, chunk_id, len(batches))
    return results


def _scan_leader(leader_row: int, all_rows: dict, cfg: Config) -> int | None:
    """Scan the leader row to find the size-chart column index.

    Returns 0-based column index (relative to COL_START) or None if not found.
    """
    links: list[str] = []
    for val in all_rows[leader_row]:
        if val is None:
            break
        links.append(val)

    if not links:
        return None

    num_links = len(links)
    # Start from O column (index 0) and scan forward.
    for si in range(0, num_links):
        url = links[si]
        p = _download(url)
        if p is None:
            continue
        try:
            if _is_size_chart(p, cfg):
                return si
        except Exception:
            pass
        finally:
            try:
                p.unlink()
            except OSError:
                pass

    return None


if __name__ == "__main__":
    main()
