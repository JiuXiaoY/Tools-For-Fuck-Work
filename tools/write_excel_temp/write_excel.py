"""临时工具:用 openpyxl 生成一个 .xlsx 文件。

用法示例:
    python tools/write_excel_temp/write_excel.py outputs/demo.xlsx
    python tools/write_excel_temp/write_excel.py outputs/demo.xlsx -s 销售数据 -r 10 -c 6
"""

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def write_excel(path: Path, sheet_name: str, rows: int, cols: int, headers: list[str] | None) -> None:
    """生成 Excel:首行为表头(加粗+填充+冻结),下面填充示例数据。"""
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if headers is None:
        headers = [f"列{i + 1}" for i in range(cols)]

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 示例数据行
    for r in range(2, rows + 2):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=f"数据{r - 1}-{c}")

    # 自动列宽(按表头文字长度)
    for col, name in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col)].width = max(len(name) * 2 + 4, 12)

    ws.freeze_panes = "A2"
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="用 openpyxl 写一个 Excel 文件")
    parser.add_argument("output", help="输出 .xlsx 文件路径")
    parser.add_argument("-s", "--sheet", default="Sheet1", help="工作表名称(默认 Sheet1)")
    parser.add_argument("-r", "--rows", type=int, default=5, help="数据行数(默认 5)")
    parser.add_argument("-c", "--cols", type=int, default=5, help="列数(默认 5)")
    parser.add_argument("-H", "--headers", nargs="*", default=None, help="自定义表头,如 -H 姓名 年龄 城市")
    args = parser.parse_args()

    if args.rows < 1 or args.cols < 1:
        print("错误:rows 和 cols 必须 >= 1", file=sys.stderr)
        sys.exit(1)

    out = Path(args.output)
    write_excel(out, args.sheet, args.rows, args.cols, args.headers)
    print(f"已生成: {out.resolve()}({args.rows} 行数据 x {args.cols} 列,sheet='{args.sheet}')")


if __name__ == "__main__":
    main()
