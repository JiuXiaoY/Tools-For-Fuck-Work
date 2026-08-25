"""临时工具:按 A 列有色单元格分区,把 prompt/finePoints/help 的内容循环写入 AZ 列。

逻辑:
- A 列有色单元格把数据分成 131 个区间,第 k 个区间 = [第 k 个有色单元格行, 第 k+1 个有色单元格行)。
- 第 k 个区间使用 help 文件第 (3k-2)~(3k) 行(共 3 行),从区间起始行开始循环写入 AZ 列,
  直到区间结束(下一个有色单元格所在行)为止。
- help 共 393 行 = 3 * 131,恰好与有色单元格数量对应。

用法:
    python tools/write_excel_temp/fill_az_from_help.py [--src 路径] [--no-backup]
"""

import argparse
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent.parent.parent
DEFAULT_SRC = BASE / "outputs" / "7.23v1.xlsx"
HELP = BASE / "y_addreoffici" / "de" / "finePoints" / "help"
AZ_COL = 52  # AZ = 第 52 列


def is_colored(cell) -> bool:
    """判断单元格是否有填充色(排除默认无色填充)。"""
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return False
    fg = fill.fgColor
    if fg is None:
        return False
    if fg.theme is not None:
        return True
    rgb = fg.rgb
    return rgb not in (None, "00000000")


def main() -> None:
    parser = argparse.ArgumentParser(description="把 help 内容按 A 列有色单元格分区循环写入 AZ 列")
    parser.add_argument("--src", default=str(DEFAULT_SRC), help="源 xlsx 路径(默认 outputs/7.21v1rewrite.xlsx)")
    parser.add_argument("--no-backup", action="store_true", help="不生成 .bak 备份")
    args = parser.parse_args()

    src = Path(args.src)
    if not src.exists():
        print(f"错误:找不到源文件 {src}", file=sys.stderr)
        sys.exit(1)

    # 1. 读取 help 内容(每 3 行一组)
    lines = [ln.strip() for ln in HELP.read_text(encoding="utf-8").splitlines()]
    lines = [ln for ln in lines if ln]
    if len(lines) % 3 != 0:
        print(f"警告:help 行数 {len(lines)} 不是 3 的倍数,最后一组不完整", file=sys.stderr)

    # 2. 备份 + 打开工作簿
    if not args.no_backup:
        bak = src.with_suffix(src.suffix + ".bak")
        shutil.copy2(src, bak)
        print(f"已备份: {bak}")

    wb = load_workbook(src)
    ws = wb.active
    print(f"工作表: {ws.title}({ws.max_row} 行 x {ws.max_column} 列)")

    # 3. 找出 A 列有色单元格行号
    colored_rows = [
        row[0].row
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1)
        if is_colored(row[0])
    ]
    if not colored_rows:
        print("错误:A 列没有找到有色单元格", file=sys.stderr)
        sys.exit(1)
    print(f"A 列有色单元格: {len(colored_rows)} 个,首行 {colored_rows[0]},末行 {colored_rows[-1]}")

    # 4. 分区循环写入 AZ 列
    bounds = colored_rows + [ws.max_row + 1]
    filled = 0
    for k, start in enumerate(colored_rows):
        end = bounds[k + 1]  # 下一个有色单元格行(不含)
        group = lines[k * 3 : k * 3 + 3]  # 本区间对应的 3 行内容
        for row in range(start, end):
            ws.cell(row=row, column=AZ_COL, value=group[(row - start) % 3])
            filled += 1

    wb.save(src)
    print(f"完成:共写入 {filled} 行 AZ 列,覆盖 {len(colored_rows)} 个区间")


if __name__ == "__main__":
    main()
