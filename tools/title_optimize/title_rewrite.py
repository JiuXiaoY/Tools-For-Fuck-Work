"""Title optimize workflow: extract → validate → optimize → write-back.

Workflow:
  1. Read outputs/{date}v1.xlsx, extract col H→origin_title, col O→origin_link
  2. Validate all three files exist and line counts match
  3. Run optimize (deepseek_web.py or run.py)
  4. Insert column after G, write optimized titles back to Excel

Usage:
    python tools/title_optimize/workflow.py
"""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from config import Config
from services.logger import get_logger

BASE = Path(__file__).resolve().parent
ORIGIN_LINK = BASE / "origin_link"
ORIGIN_TITLE = BASE / "origin_title"
OPTIMIZE_TITLE = BASE / "optimize_title"

_log = get_logger("title_workflow")


def main() -> None:
    cfg = Config()

    # ── 1. Find source Excel ──
    date_str = _resolve_date(cfg)
    out_dir = Path(__file__).resolve().parent.parent.parent / cfg.out_dir
    src = out_dir / f"{date_str}v1.xlsx"
    if not src.exists():
        _log.error("Source not found: %s", src)
        return
    _log.info("Source: %s", src.name)

    # ── 2. Extract titles and links ──
    wb = openpyxl.load_workbook(src)
    ws = wb.active

    titles: list[str] = []
    links: list[str] = []
    from services import cell_has_fill

    for r in range(1, ws.max_row + 1):
        if not cell_has_fill(ws.cell(row=r, column=1)):
            continue
        h_val = ws.cell(row=r, column=8).value
        o_val = ws.cell(row=r, column=15).value
        if h_val is None or str(h_val).strip() == "":
            _log.error("Row %d: col H is empty", r)
            wb.close()
            return
        if o_val is None or str(o_val).strip() == "":
            _log.error("Row %d: col O is empty", r)
            wb.close()
            return
        titles.append(str(h_val).strip())
        links.append(str(o_val).strip())

    wb.close()
    _log.info("Extracted %d titles + %d links", len(titles), len(links))

    if not titles:
        _log.error("No data extracted (no rows with col A fill)")
        return

    # ── 3. Write to files ──
    ORIGIN_TITLE.write_text("\n".join(titles), encoding="utf-8")
    ORIGIN_LINK.write_text("\n".join(links), encoding="utf-8")
    _log.info("Written: origin_title (%d) + origin_link (%d)", len(titles), len(links))

    # ── 4. Check optimize_title ──
    if OPTIMIZE_TITLE.exists():
        opt_content = OPTIMIZE_TITLE.read_text(encoding="utf-8").strip()
    else:
        opt_content = ""

    opt_lines = [l.strip() for l in opt_content.splitlines() if l.strip()]

    if opt_lines and len(opt_lines) == len(titles) and opt_lines != titles:
        # Already optimized — skip to write-back
        _log.info("optimize_title already has %d optimized lines, skipping optimization", len(opt_lines))
    else:
        # Need to run optimization
        _log.info("optimize_title empty or outdated, running optimization...")
        # Seed with origin titles so run.py has input
        OPTIMIZE_TITLE.write_text("")  # clear for fresh run

        # if mode == "api":
        #     script = BASE / "run.py"
        # else:
        script = BASE / "deepseek_web.py"

        _log.info("Running: %s", script.name)
        result = subprocess.run(
            [sys.executable, str(script)],
            cwd=str(BASE),
        )
        if result.returncode != 0:
            _log.error("Optimization failed with code %d", result.returncode)
            return

        # Re-read after optimization
        opt_content = OPTIMIZE_TITLE.read_text(encoding="utf-8").strip()
        opt_lines = [l.strip() for l in opt_content.splitlines() if l.strip()]

    # Validate final
    if not opt_lines:
        _log.error("optimize_title is empty after optimization")
        return
    if len(opt_lines) != len(titles):
        _log.error("Count mismatch: origin=%d, optimize=%d", len(titles), len(opt_lines))
        return
    _log.info("Optimization complete: %d titles", len(opt_lines))

    # ── 6. Write back to Excel ──
    wb = openpyxl.load_workbook(src)
    ws = wb.active

    # Insert one column after G (col 7) → new column at position 8
    ws.insert_cols(8)
    _log.info("Inserted new column at position 8 (after G)")

    # Write optimized titles to the new column H (8), matching original rows
    opt_idx = 0
    written = 0
    for r in range(1, ws.max_row + 1):
        if not cell_has_fill(ws.cell(row=r, column=1)):
            continue
        ws.cell(row=r, column=8).value = opt_lines[opt_idx]
        opt_idx += 1
        written += 1

    wb.save(src)
    wb.close()
    _log.info("Written %d optimized titles to column H (8)", written)
    _log.info("Done: %s", src.name)


def _resolve_date(cfg: Config) -> str:
    raw = cfg.date_override.strip()
    if raw and len(raw) == 6:
        try:
            month = int(raw[2:4])
            day = int(raw[4:6])
            return f"{month}.{day}"
        except ValueError:
            pass
    from datetime import datetime
    today = datetime.now()
    return f"{today.month}.{today.day}"


if __name__ == "__main__":
    main()
