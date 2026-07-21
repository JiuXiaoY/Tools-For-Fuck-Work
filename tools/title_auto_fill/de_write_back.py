"""Write final_de_title to column 4, then fill blanks with cell-above formula.

Usage:
    python tools/title_auto_fill/write_back.py
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from config import Config
from services.logger import get_logger

BASE = Path(__file__).resolve().parent
FINAL_TITLES = BASE / "final_de_title"
_log = get_logger("write_back")


def main() -> None:
    cfg = Config()
    date_str = _resolve_date(cfg)
    src = Path(__file__).resolve().parent.parent.parent / cfg.out_dir / f"{date_str}v1.xlsx"

    if not src.exists():
        _log.error("Source not found: %s", src)
        return

    if not FINAL_TITLES.exists():
        _log.error("final_de_title not found: %s", FINAL_TITLES)
        return

    titles = [l.strip() for l in FINAL_TITLES.read_text(encoding="utf-8").strip().splitlines()]
    _log.info("Loaded %d titles from final_de_title", len(titles))

    from services import cell_has_fill

    wb = openpyxl.load_workbook(src)
    ws = wb.active

    # ── Write titles to column 4 ──
    idx = 0
    written = 0
    for r in range(1, ws.max_row + 1):
        if cell_has_fill(ws.cell(row=r, column=1)):
            if idx < len(titles):
                ws.cell(row=r, column=4).value = titles[idx]
                idx += 1
                written += 1

    _log.info("Written %d titles to column 4", written)

    # ── Fill blank cells with =D{row-1} ──
    filled = 0
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=4)
        if cell.value is None or str(cell.value).strip() == "":
            cell.value = f"=D{r - 1}"
            filled += 1

    wb.save(src)
    wb.close()
    _log.info("Filled %d blank cells with =D{above}", filled)
    _log.info("Done: %s", src.name)


def _resolve_date(cfg: Config) -> str:
    raw = cfg.date_override.strip()
    if raw and len(raw) == 6:
        try:
            month = int(raw[2:4])
            day = int(raw[4:6])
            return f"{month}.{day}"
        except ValueError:
            pass
    from datetime import datetime
    today = datetime.now()
    return f"{today.month}.{today.day}"


if __name__ == "__main__":
    main()
