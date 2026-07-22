"""Export first two columns from output Excel to a new sheet.

Output: outputs/{M}.{D} - N.xlsx
Headers: 系统SKU/捆绑SKU | 平台SKU

Usage:
    python tools/export_sku.py
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from config import Config
from services.logger import get_logger

_log = get_logger("export_sku")


def main() -> None:
    cfg = Config()
    date_str = _resolve_date(cfg)
    src = Path(__file__).resolve().parent.parent.parent / cfg.out_dir / f"{date_str}v1.xlsx"

    if not src.exists():
        _log.error("Source not found: %s", src)
        return

    dst = Path(__file__).resolve().parent.parent.parent / cfg.out_dir / f"{date_str} - N.xlsx"

    # Read source
    wb_src = openpyxl.load_workbook(src, data_only=True)
    ws_src = wb_src.active

    # Create output
    wb_dst = openpyxl.Workbook()
    ws_dst = wb_dst.active
    ws_dst.title = "SKU"

    # Headers
    ws_dst.cell(row=1, column=1).value = "系统SKU/捆绑SKU"
    ws_dst.cell(row=1, column=2).value = "平台SKU"

    # Copy values only (no formatting)
    row_out = 2
    for r in range(1, ws_src.max_row + 1):
        val_a = ws_src.cell(row=r, column=1).value
        val_b = ws_src.cell(row=r, column=2).value
        if val_a is None and val_b is None:
            continue
        ws_dst.cell(row=row_out, column=1).value = val_a
        ws_dst.cell(row=row_out, column=2).value = val_b
        row_out += 1

    # Formatting
    ws_dst.column_dimensions['A'].width = 20
    ws_dst.column_dimensions['B'].width = 20
    for r in range(1, row_out):
        ws_dst.row_dimensions[r].height = 13.5

    wb_src.close()
    wb_dst.save(dst)
    wb_dst.close()
    _log.info("Exported %d rows → %s", row_out - 2, dst.name)


def _resolve_date(cfg: Config) -> str:
    raw = cfg.date_override.strip()
    if raw and len(raw) == 6:
        try:
            month = int(raw[2:4])
            day = int(raw[4:6])
            return f"{month}.{day}"
        except ValueError:
            pass
    from datetime import datetime
    today = datetime.now()
    return f"{today.month}.{today.day}"


if __name__ == "__main__":
    main()
