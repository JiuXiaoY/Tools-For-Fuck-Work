"""Color-size mapping for Excel: extract → process → write-back.

Workflow:
  1. Read outputs/{date}v1.xlsx, col J(10) + K(11)
  2. Write to check_.txt (tab-separated, empty lines as group separators)
  3. Run process.py to handle duplicates + strip size
  4. Read processed check_.txt → write back to column J

Usage:
    python tools/color_size_deal/main.py
"""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from config import Config
from services.logger import get_logger

BASE = Path(__file__).resolve().parent
CHECK_TXT = BASE / "check_.txt"
PROCESS_PY = BASE / "process.py"

_log = get_logger("color_size_main")


def main() -> None:
    cfg = Config()

    # ── 1. Find source Excel ──
    date_str = _resolve_date(cfg)
    out_dir = Path(__file__).resolve().parent.parent.parent / cfg.out_dir
    src = out_dir / f"{date_str}v1.xlsx"
    if not src.exists():
        _log.error("Source not found: %s", src)
        return
    _log.info("Source: %s", src.name)

    # ── 2. Extract J + K → check_.txt ──
    wb = openpyxl.load_workbook(src)
    ws = wb.active

    lines: list[str] = []
    empty_count = 0
    data_count = 0

    for r in range(1, ws.max_row + 1):
        j_val = ws.cell(row=r, column=10).value
        k_val = ws.cell(row=r, column=11).value

        j_str = str(j_val).strip() if j_val is not None else ""
        k_str = str(k_val).strip() if k_val is not None else ""

        if not j_str or not k_str:
            lines.append("")  # empty line as group separator
            empty_count += 1
        else:
            lines.append(f"{j_str}\t{k_str}")
            data_count += 1

    wb.close()
    CHECK_TXT.write_text("\n".join(lines), encoding="utf-8")
    _log.info("Extracted: %d data lines + %d separators → %s", data_count, empty_count, CHECK_TXT.name)

    # ── 3. Run process.py ──
    _log.info("Running: %s", PROCESS_PY.name)
    result = subprocess.run(
        [sys.executable, str(PROCESS_PY)],
        cwd=str(BASE),
    )
    if result.returncode != 0:
        _log.error("process.py failed with code %d", result.returncode)
        return

    # ── 4. Read processed check_.txt → write back to col J ──
    processed_lines = CHECK_TXT.read_text(encoding="utf-8").split("\n")
    _log.info("Processing write-back: %d lines", len(processed_lines))

    wb = openpyxl.load_workbook(src)
    ws = wb.active

    # First pass: count data rows to validate
    proc_idx = 0
    written = 0
    skipped = 0

    for r in range(1, ws.max_row + 1):
        if proc_idx >= len(processed_lines):
            break

        new_val = processed_lines[proc_idx].strip()
        proc_idx += 1

        # Get current value
        old_val = ws.cell(row=r, column=10).value
        old_str = str(old_val).strip() if old_val is not None else ""

        if new_val == "":  # separator
            ws.cell(row=r, column=10).value = None
            written += 1
        elif new_val != old_str:  # only write if changed
            ws.cell(row=r, column=10).value = new_val
            written += 1
        else:
            skipped += 1

    wb.save(src)
    wb.close()
    _log.info("Write-back: %d written, %d unchanged → %s", written, skipped, src.name)
    _log.info("Done.")


def _resolve_date(cfg: Config) -> str:
    raw = cfg.date_override.strip()
    if raw and len(raw) == 6:
        try:
            month = int(raw[2:4])
            day = int(raw[4:6])
            return f"{month}.{day}"
        except ValueError:
            pass
    from datetime import datetime
    today = datetime.now()
    return f"{today.month}.{today.day}"


if __name__ == "__main__":
    main()
