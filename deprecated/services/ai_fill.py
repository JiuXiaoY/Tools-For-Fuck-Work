"""AI fill: Gemini-powered column 4 generation.

Assembles system instruction from:
  instructions/{category}.md — rules, examples, keyword references
"""

from __future__ import annotations

import re
from pathlib import Path

from google import genai
from openpyxl import load_workbook

from config import Config

_PROMPT_DIR = Path(__file__).resolve().parent.parent / "prompt_de"
_INCLUDE_RE = re.compile(r"\{@include\s+(.+?)\}")

_CAT_KEY = {"女士上衣": "womens_tops", "男士上衣": "mens_tops",
           "女士裤子": "womens_pants", "男士裤子": "mens_pants",
           "女士套装": "womens_suits", "男士套装": "mens_suits",
           "女士连衣裙": "womens_dresses", "女士连体衣": "womens_onesies"}


def _resolve(text: str, visited: set[str] | None = None) -> str:
    if visited is None:
        visited = set()

    def _read(path: str) -> str:
        full = (_PROMPT_DIR / path).resolve()
        key = str(full)
        if key in visited:
            return ""
        visited.add(key)
        return _resolve(full.read_text(encoding="utf-8"), visited)

    text = _INCLUDE_RE.sub(lambda m: _read(m.group(1).strip()), text)
    return text


def _load_system_instruction(config: Config) -> str:
    cat = config.gemini_category
    path = _PROMPT_DIR / "instructions" / f"{cat}.md"
    if not path.exists():
        raise FileNotFoundError(f"Instructions not found: {path}")
    return _resolve(path.read_text(encoding="utf-8"))


def fill_column_4(xlsx_path: str | Path, config: Config | None = None) -> int:
    """For rows where col A has fill, read col 8, chat with Gemini, write col 4."""
    if config is None:
        config = Config()
    if not config.gemini_api_key:
        raise ValueError("gemini_api_key is empty")

    system = _load_system_instruction(config)
    client = genai.Client(api_key=config.gemini_api_key)
    chat = client.chats.create(
        model=config.gemini_model,
        config={"system_instruction": system},
    )

    wb = load_workbook(xlsx_path)
    ws = wb.active
    from services import cell_has_fill

    filled = 0
    for r in range(1, ws.max_row + 1):
        if not cell_has_fill(ws.cell(row=r, column=1)):
            continue
        source = ws.cell(row=r, column=8).value
        if source is None or str(source).strip() == "":
            continue
        text = str(source).strip()
        try:
            resp = chat.send_message(text)
            ws.cell(row=r, column=4).value = resp.text.strip()
            filled += 1
        except Exception as exc:
            ws.cell(row=r, column=4).value = f"[ERR:{exc}]"

    wb.save(xlsx_path)
    return filled
