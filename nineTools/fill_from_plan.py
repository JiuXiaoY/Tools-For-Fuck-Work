# -*- coding: utf-8 -*-
"""
fill_from_plan —— 按「填充计划(plan)」把数据写入模板副本。

只读模板，写新文件；所有"填哪列、填哪些值、顺序"由 plan JSON 描述，本程序不硬编码任何列。

═══════════════════════════════════════════════════════════════════════════════
Plan JSON 格式
═══════════════════════════════════════════════════════════════════════════════
{
  "template_file": "prompt_fr/final_init_template/coat_template_Eva.xlsm",  # 只读模板
  "output_file":   "outputs/filled.xlsm",                                   # 输出副本(可选,默认<模板>_filled.xlsm)
  "data_start_row": 7,           # 模板数据起始行。组内行号以"数据区第1行=1"计, 实际行=组行号+data_start_row-1
  "col_scope": [1, 2, 4, 6, 12, 17, 21, 40, 133],   # 允许填写的列(非连续集合); 不在其中的列绝不写
  "groups": {                    # 分组: 组名 -> "父体起始行 & 结束行"(窗口坐标)
    "group_1": "1 & 99",
    "group_2": "100 & 199"
  },
  "cycle_threshold": null,       # 循环阈值: 当 0<数据条数<阈值 时循环铺满; null=任何"少于行数"都循环
  "data": {                      # 每组每列数据: 组名 -> { 列号(str) : [值...] }
    "group_1": { "12": ["<可选值1>", "<可选值2>"], "133": [...] },
    "group_2": { "12": [...] }
  }
}

每列填充规则(设该组数据行数=n, 该列数据条数=m):
  m == 0           -> 不填
  m == n           -> 按顺序填(父+子, 从组内第一行开始)
  m == n - 1       -> 只填子体(跳过父体那一行, 从第二行开始)
  0 < m < 阈值      -> 循环铺满整组行
  其它(m>n 或阈值禁用循环时的中间值) -> mismatch: 不写, 记入报告
═══════════════════════════════════════════════════════════════════════════════

用法:
    python nineTools/fill_from_plan.py <plan.json> [--report 报告路径]
"""

import argparse
import json
import os
import sys
from pathlib import Path

import openpyxl


# --------------------------------------------------------------------------- #
# 填充规则
# --------------------------------------------------------------------------- #
def apply_column(
    ws,
    group_start_actual: int,
    group_end_actual: int,
    column: int,
    values,
    cycle_threshold,
):
    """
    将 values 按给定规则填入 group 的某列。

    返回: (mode, filled_count, note)
        mode: 'none'|'sequential'|'children_only'|'cycle'|'mismatch'
    """
    n = group_end_actual - group_start_actual + 1   # 该组数据行数
    m = len(values)

    def _cycle_allowed():
        return cycle_threshold is None or m < cycle_threshold

    if m == 0:
        return "none", 0, "数据为空, 不填"
    if m == n:
        target = list(range(group_start_actual, group_end_actual + 1))
        mode = "sequential"
    elif m == n - 1:
        target = list(range(group_start_actual + 1, group_end_actual + 1))  # 跳过父体
        mode = "children_only"
    elif m < n and _cycle_allowed():
        target = list(range(group_start_actual, group_end_actual + 1))
        mode = "cycle"
    else:
        # m>n，或阈值禁用循环但仍少于行数 -> 数据与组行数不匹配
        return "mismatch", 0, f"数据条数({m})与组行数({n})不匹配"

    filled = 0
    for i, row in enumerate(target):
        if i >= len(values):
            break  # cycle 分支中 target 长于 values, 用余数取
        value = values[i % m] if mode == "cycle" else values[i]
        ws.cell(row=row, column=column).value = value
        filled += 1
    note = {
        "none": "不填",
        "sequential": f"顺序填写 {filled} 格(父+子)",
        "children_only": f"只填子体 {filled} 格(跳过父体)",
        "cycle": f"循环填写 {filled} 格(模式长度 {m})",
    }.get(mode, "")
    return mode, filled, note


# --------------------------------------------------------------------------- #
# 主流程
# --------------------------------------------------------------------------- #
def main():
    parser = argparse.ArgumentParser(
        description="按填充计划(plan.json)把数据写入模板副本(只读模板)。"
    )
    parser.add_argument("plan", help="填充计划 JSON 文件路径")
    parser.add_argument("--report", default=None, help="填充报告 JSON 输出路径(可选)")
    parser.add_argument(
        "--strict-scope",
        action="store_true",
        help="开启后: data 中出现的列若不在 col_scope 内则报错退出(默认跳过并警告)",
    )
    args = parser.parse_args()

    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    # ---- 读 plan ----
    try:
        plan = json.loads(Path(args.plan).read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001
        print(f"无法读取 plan 文件 {args.plan}: {exc}")
        sys.exit(1)

    template_file = plan.get("template_file")
    if not template_file:
        print("plan 缺少 template_file")
        sys.exit(2)
    out_file = plan.get("output_file") or _default_output(template_file)
    data_start_row = int(plan.get("data_start_row") or 7)
    col_scope = set(plan.get("col_scope") or [])
    groups = plan.get("groups") or {}
    data = plan.get("data") or {}
    cycle_threshold = plan.get("cycle_threshold")

    # ---- 打开模板(只读), 复制写新文件 ----
    if not os.path.exists(template_file):
        print(f"模板文件不存在: {template_file}")
        sys.exit(3)
    wb = openpyxl.load_workbook(template_file, keep_vba=True, data_only=False)
    ws_name = "Modèle" if "Modèle" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[ws_name]

    offset = data_start_row - 1  # 组内行号 -> 实际模板行

    report_entries = []
    total_filled = 0
    out_of_scope = []

    # ---- 遍历每组每列 ----
    for gname, spec in groups.items():
        spec = str(spec).strip()
        if "&" not in spec:
            print(f"[skip] 组 {gname} 的规格无效: {spec!r}")
            continue
        try:
            start_w, end_w = map(int, spec.split("&"))
        except ValueError:
            print(f"[skip] 组 {gname} 的规格无效: {spec!r}")
            continue
        start_actual = start_w + offset
        end_actual = end_w + offset
        if start_actual < 1 or end_actual < start_actual:
            print(f"[skip] 组 {gname} 换算后行范围非法: {start_actual}..{end_actual}")
            continue

        col_data = data.get(gname) or {}
        for col_str, values in col_data.items():
            try:
                col = int(col_str)
            except ValueError:
                print(f"[skip] 组 {gname} 列号非法: {col_str!r}")
                continue
            if col not in col_scope:
                out_of_scope.append((gname, col))
                if args.strict_scope:
                    print(f"[error] 列 {col} 不在 col_scope 内(组 {gname}), 开启 --strict-scope 报错退出。")
                    sys.exit(4)
                print(f"[warn] 列 {col} 不在 col_scope 内(组 {gname}), 跳过不填。")
                continue
            mode, filled, note = apply_column(
                ws, start_actual, end_actual, col, values, cycle_threshold
            )
            total_filled += filled
            report_entries.append(
                {
                    "group": gname,
                    "rows": [start_actual, end_actual],
                    "column": col,
                    "mode": mode,
                    "filled": filled,
                    "count": len(values),
                    "note": note,
                }
            )
            print(
                f"  {gname:8} 列{col:<4} 模式={mode:<13} 填={filled:<3} 条数={len(values)}  {note}"
            )

    # ---- 保存副本 ----
    out_dir = os.path.dirname(os.path.abspath(out_file))
    os.makedirs(out_dir, exist_ok=True)
    wb.save(out_file)

    # ---- 报告 ----
    report = {
        "template_file": template_file,
        "output_file": out_file,
        "data_start_row": data_start_row,
        "col_scope": sorted(col_scope),
        "groups": groups,
        "cycle_threshold": cycle_threshold,
        "total_filled": total_filled,
        "entries": report_entries,
        "out_of_scope_skipped": out_of_scope,
    }
    print(f"\n已写出副本: {out_file}")
    print(f"共填写 {total_filled} 个单元格。")
    mismatches = [e for e in report_entries if e["mode"] == "mismatch"]
    if mismatches:
        print(f"注意: 有 {len(mismatches)} 个 group/列 因数据与行数不匹配而未填写, 详见报告。")
    if out_of_scope:
        print(f"注意: 有 {len(out_of_scope)} 个列因不在 col_scope 内被跳过。")

    if args.report:
        Path(args.report).write_text(
            json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(f"已写出报告: {args.report}")


def _default_output(template_file: str) -> str:
    base = os.path.splitext(os.path.basename(template_file))[0]
    return os.path.join(os.path.dirname(os.path.abspath(template_file)), f"{base}_filled.xlsm")


if __name__ == "__main__":
    main()
